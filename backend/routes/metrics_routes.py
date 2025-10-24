# routes/metrics_routes.py
from __future__ import annotations

from flask import Blueprint, jsonify, current_app, request, session as _session
from pathlib import Path
import logging, os, time, math, re, warnings
from datetime import datetime, timedelta, date

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from backend.extensions import db
from sqlalchemy.dialects.sqlite import insert as sqlite_insert  # swap for other DBs
from sqlalchemy import and_, or_, func

# --- Models
from backend.models import (
    SharePointConnection,
    PortfolioPeriodMetric,
    MarketPrice,
    User,
    Investor,
    PortfolioInvestmentValue,
    Investment
)
# Optional snapshot models (used by investor-overview DB-first)
try:
    from backend.models_snapshot import WorkbookSnapshot, InvestorBalance, InvestorPeriodBalance
except Exception:
    WorkbookSnapshot = None
    InvestorBalance = None

# --- Graph helpers
from backend.routes.auth_ms_routes import get_session_bearer  # user (delegated) bearer
from backend.routes.sharepoint_excel_routes import _bearer_from_request  # request-bearer fallback
from backend.graph_excel_live import create_session, close_session, used_range_values
try:
    from backend.graph_excel_live import write_range_value
except Exception:
    write_range_value = None
from backend.graph_sharepoint import (
    list_worksheets,
    open_excel_by_share_url,
    read_range as sp_read_range,
)

# --- Extra services
try:
    from services.market_store import upsert_history
except Exception:
    upsert_history = None

logging.basicConfig(level=logging.INFO)

# =========================
#   Blueprints
# =========================
metrics_bp = Blueprint("metrics_bp", __name__, url_prefix="/api/metrics")
market_bp  = Blueprint("market",     __name__, url_prefix="/api/market")

# =========================
#   Small caches
# =========================
_OVERVIEW_CACHE = {}
_APP_BEARER_CACHE = {"token": None, "exp": 0, "err_status": None, "err_text": None, "ts": 0}

# =========================
#   Config helpers
# =========================
def _cfg(key: str, default=None):
    """Read from Flask config, falling back to Config class defaults."""
    if current_app and getattr(current_app, "config", None):
        if key in current_app.config and current_app.config[key] is not None:
            return current_app.config[key]
    try:
        from config import Config
        return getattr(Config, key, default)
    except Exception:
        return default

# =========================
#   Path helpers
# =========================
def _norm(s: str) -> str:
    return re.sub(r"[ \-_.()]+", "", s or "").lower()

def _uploads_dir() -> Path:
    """
    Prefer UPLOAD_ROOT (from app.py), then common fallbacks.
    """
    root = Path(current_app.root_path)
    candidates = []
    up_root = _cfg("UPLOAD_ROOT")
    if up_root:
        candidates.append(Path(up_root))
    up_folder = current_app.config.get("UPLOAD_FOLDER")
    if up_folder:
        p = Path(up_folder)
        candidates.append(p if p.is_absolute() else (root / p))
    candidates += [
        root / "uploads",
        root.parent / "uploads",
        Path.cwd() / "uploads",
        root / "static" / "uploads",
    ]
    for c in candidates:
        try:
            c.mkdir(parents=True, exist_ok=True)
        except Exception:
            pass
    for c in candidates:
        if c.exists():
            return c.resolve()
    fb = root / "uploads"
    fb.mkdir(parents=True, exist_ok=True)
    return fb.resolve()

def _find_best_xlsm(uploads: Path, filename: str | None) -> Path:
    files = sorted(uploads.glob("*.xlsm"))
    if filename:
        p = Path(filename)
        if p.is_absolute() and p.exists():
            return p.resolve()
        p2 = (uploads / filename).resolve()
        if p2.exists():
            return p2
        want = _norm(filename)
        for f in files:
            if _norm(f.name) == want or want in _norm(f.name):
                return f.resolve()
        raise FileNotFoundError(f"Workbook not found: {filename} in {uploads}")
    if not files:
        raise FileNotFoundError(f"No .xlsm files in {uploads}")
    elpis = [f for f in files if f.name.lower().startswith("elpis")]
    return (elpis[0] if elpis else files[0]).resolve()

def _discover_xlsm(file: str | None, path: str | None):
    """
    Strong discovery for local .xlsm.
    Returns (Path, searched_dirs[list[str]])
    """
    searched = []

    if path:
        p = Path(path)
        if not p.is_absolute():
            p = (Path(current_app.root_path) / p).resolve()
        if p.is_file() and p.suffix.lower() == ".xlsm":
            return p, searched
        searched.append(str(p))

    default_file = _cfg("DEFAULT_WORKBOOK_FILE")
    if default_file:
        pdef = Path(default_file)
        if not pdef.is_absolute():
            pdef = (Path(current_app.root_path) / pdef).resolve()
        if pdef.is_file() and pdef.suffix.lower() == ".xlsm":
            return pdef, searched
        searched.append(str(pdef))

    try:
        up = _uploads_dir()
        if file:
            return _find_best_xlsm(up, file), searched
    except Exception:
        pass

    roots = [
        _uploads_dir(),
        Path(current_app.root_path),
        Path(current_app.root_path).parent,
        Path.cwd(),
        Path(current_app.root_path) / "static" / "uploads",
    ]
    for r in roots:
        searched.append(str(r))
        picks = sorted(r.glob("*.xlsm"))
        if picks:
            elpis = [f for f in picks if f.name.lower().startswith("elpis")]
            return (elpis[0] if elpis else picks[0]).resolve(), searched

    raise FileNotFoundError(f"No .xlsm found. Searched: {searched}")

# =========================
#   Bearer helpers
# =========================
def _app_token_last_error():
    return _APP_BEARER_CACHE.get("err_status"), _APP_BEARER_CACHE.get("err_text"), _APP_BEARER_CACHE.get("ts")

def _get_app_bearer() -> str | None:
    """
    Client-credentials token with one-shot Conditional Access claims retry.
    """
    allow = str(_cfg("GRAPH_ALLOW_APP_FALLBACK", "true")).lower() in ("1", "true", "yes")
    if not allow:
        return None

    now = time.time()
    if _APP_BEARER_CACHE["token"] and _APP_BEARER_CACHE["exp"] - 60 > now:
        return _APP_BEARER_CACHE["token"]

    tenant = _cfg("AZURE_TENANT_ID")
    cid    = _cfg("AZURE_CLIENT_ID")
    secret = _cfg("AZURE_CLIENT_SECRET")
    scope  = _cfg("GRAPH_SCOPES", "https://graph.microsoft.com/.default")

    if not (tenant and cid and secret):
        _APP_BEARER_CACHE.update({"err_status": "config_missing", "err_text": "Missing AZURE_TENANT_ID/CLIENT_ID/CLIENT_SECRET", "ts": now})
        return None

    url = f"https://login.microsoftonline.com/{tenant}/oauth2/v2.0/token"
    base = {"client_id": cid, "client_secret": secret, "grant_type": "client_credentials", "scope": scope}
    try:
        def _post(payload):
            import requests
            return requests.post(url, data=payload, timeout=15)

        resp = _post(base)

        # Conditional Access claims
        if resp.status_code == 400:
            try:
                jerr = resp.json()
                if jerr.get("error") == "invalid_grant" and "claims" in jerr:
                    data_with_claims = dict(base)
                    data_with_claims["claims"] = jerr["claims"]
                    resp = _post(data_with_claims)
            except Exception:
                pass

        if resp.status_code != 200:
            txt = (resp.text or "")[:800]
            _APP_BEARER_CACHE.update({"err_status": resp.status_code, "err_text": txt, "ts": now})
            return None
        j = resp.json()
        tok = j.get("access_token")
        exp = now + int(j.get("expires_in", 3600))
        if tok:
            _APP_BEARER_CACHE.update({"token": tok, "exp": exp, "err_status": None, "err_text": None, "ts": now})
            return tok
        _APP_BEARER_CACHE.update({"err_status": "no_token", "err_text": str(j)[:800], "ts": now})
        return None
    except Exception as e:
        _APP_BEARER_CACHE.update({"err_status": "exception", "err_text": f"{type(e).__name__}: {e}", "ts": now})
        return None

def _resolve_bearer() -> str | None:
    # Prefer live delegated token from session; then request-bearer; then app token.
    return get_session_bearer() or _bearer_from_request() or _get_app_bearer()

def _unauthorized_response_no_bearer(context_hint: str, searched_uploads: str | None = None):
    status, text, ts = _app_token_last_error()
    detail = ""
    http_code = 401
    if text:
        if "AADSTS53003" in text:
            http_code = 403
            detail = ("Conditional Access is blocking or challenging app tokens (AADSTS53003). "
                      "We retried with the claims challenge. If it still fails, exclude this app "
                      "from the policy or allow workload identities to satisfy the policy.")
        else:
            detail = f"Last app-token error: {status} {text[:300]}"
    msg = {
        "error": f"Unauthorized: no Microsoft token available for {context_hint}.",
        "hint": "Sign in with Microsoft (delegated), or enable client-credentials app tokens.",
        "conditional_access": detail or None,
        "searched_uploads": searched_uploads,
    }
    resp = jsonify({k: v for k, v in msg.items() if v is not None}); resp.headers["Cache-Control"] = "no-store"
    return resp, http_code

# =========================
#   Sheet name helpers
# =========================
_WS_NORMALIZE_RE = re.compile(r"[\s\u00A0\-\_\.\(\)\[\]\{\}\\+]+", re.UNICODE)
def _normalize_sheet_name(s: str) -> str:
    s = (s or "").strip().replace("\u00A0", " ")
    return _WS_NORMALIZE_RE.sub("", s).lower()

def _sheet_candidates(name: str) -> list[str]:
    base = (name or "").strip()
    variants = {
        base,
        base.replace("(", " (").replace("  ", " ").strip(),
        re.sub(r"\s*\(", " (", base).strip(),
        re.sub(r"\s+", "", base),
        base.replace("+", " "),
    } if base else set()
    return [v for v in variants if v]

def _resolve_remote_sheet_name(drive_id: str, item_id: str, wanted: str, bearer: str):
    try:
        sheets = list_worksheets(drive_id, item_id, bearer, tenant_id=None) or []
        names = []
        for s in sheets:
            nm = s.get("name") if isinstance(s, dict) else str(s)
            if nm:
                names.append(nm)
        if not names:
            return None, []
        want = _normalize_sheet_name(wanted)
        for n in names:
            if _normalize_sheet_name(n) == want:
                return n, names
        for n in names:
            if want in _normalize_sheet_name(n):
                return n, names
        for n in names:
            if n.lower().startswith((wanted or "").strip().lower()):
                return n, names
        return None, names
    except Exception as e:
        current_app.logger.exception("list_worksheets failed: %s", e)
        return None, []

# =========================
#   Value / date helpers
# =========================
def _to_float(v):
    if v in (None, "", "—", "-", "–"):
        return math.nan
    s = str(v).strip().replace(",", "").replace("$", "")
    if s.startswith("(") and s.endswith(")"):
        s = "-" + s[1:-1]
    try:
        return float(s)
    except Exception:
        return math.nan

def _parse_excel_date(val):
    if isinstance(val, datetime):
        return val
    if isinstance(val, (int, float)):
        if 20000 < float(val) < 90000:
            try:
                return datetime(1899, 12, 30) + timedelta(days=float(val))
            except Exception:
                pass
    if isinstance(val, str):
        s = val.strip()
        for fmt in ("%m/%d/%Y", "%d/%m/%Y", "%Y-%m-%d", "%m/%d/%y", "%d-%b-%Y"):
            try:
                return datetime.strptime(s, fmt)
            except Exception:
                continue
    return None

def _is_sane_date(dt: datetime | None) -> bool:
    if not dt:
        return False
    lo = datetime(2000, 1, 1)
    hi = datetime.utcnow() + timedelta(days=31)
    return lo <= dt <= hi

def _normhdr(x: str) -> str:
    s = (x or "").lower().replace("\u00a0", " ").replace("\u2011", "-")
    return re.sub(r"[^a-z0-9]+", "", s)

_DATE_ALIASES_N = (
    "asofdate", "asof", "date", "period", "month",
    "valuationdate", "navdate", "statementdate", "pricingdate",
)
_ENDING_ALIASES_N = ("ytdendingbalance", "endingbalance", "closingbalance", "currentvalue")
_BEGIN_ALIASES_N  = (
    "currentperiodbegbalance", "currentperiodbeginningbalance",
    "beginningbalance", "openingbalance", "openingnav"
)

# =========================
#   Openpyxl scan helpers
# =========================
def _row_texts(ws, r, max_col):
    out = []
    for c in range(1, max_col + 1):
        v = ws.cell(row=r, column=c).value
        out.append("" if v is None else str(v).strip().lower())
    return out

def _scan_total_row_uploads(ws, start_row: int, max_scan_rows: int = 150, label_cols: int = 12) -> int | None:
    end_row = min(ws.max_row, start_row + max_scan_rows)
    for r in range(start_row, end_row + 1):
        for c in range(1, min(label_cols, ws.max_column) + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and "total" in v.lower():
                return r
    return None

def _sum_contiguous_block_uploads(ws, start_row: int, col: int, blank_run_stop: int = 3) -> float:
    blanks = 0
    total = 0.0
    r = start_row
    while r <= ws.max_row:
        v = ws.cell(row=r, column=col).value
        num = _to_float(v)
        if math.isnan(num):
            blanks += 1
            if blanks >= blank_run_stop:
                break
        else:
            blanks = 0
            total += float(num)
        r += 1
    return float(total)

# =========================
#   Header/column detection
# =========================
def _find_header_row_and_cols(ws, max_scan_rows=None, max_scan_cols=None):
    max_scan_rows = ws.max_row if max_scan_rows is None else min(max_scan_rows, ws.max_row)
    max_scan_cols = ws.max_column if max_scan_cols is None else min(max_scan_cols, ws.max_column)

    asof_row = asof_col = None
    beg_row, end_row = None, None
    beg_cols, end_cols = [], []

    scan_ranges = [(1, min(max_scan_rows, 50))]
    if max_scan_rows > 50:
        scan_ranges.append((51, max_scan_rows))

    for start_r, stop_r in scan_ranges:
        for r in range(start_r, stop_r + 1):
            raw = _row_texts(ws, r, max_scan_cols)
            texts = [_normhdr(v) for v in raw]

            if asof_col is None:
                for i, t in enumerate(texts):
                    if any(alias in t for alias in _DATE_ALIASES_N):
                        asof_row, asof_col = r, i + 1
                        break

            for i, t in enumerate(texts):
                if any(alias in t for alias in _ENDING_ALIASES_N):
                    end_row = r
                    end_cols.append(i + 1)

            for i, t in enumerate(texts):
                if any(alias in t for alias in _BEGIN_ALIASES_N):
                    beg_row = r
                    beg_cols.append(i + 1)

            if end_cols and asof_col:
                return (asof_row, asof_col, beg_row, beg_cols, end_row, end_cols)

    return (asof_row, asof_col, beg_row, beg_cols, end_row, end_cols)

def _coerce_dates_series(col: pd.Series) -> pd.Series:
    out = pd.Series(pd.NaT, index=col.index, dtype="datetime64[ns]")

    # Excel serials
    n = pd.to_numeric(col, errors="coerce")
    mask_num = n.between(20000, 90000)
    if mask_num.any():
        parsed = pd.to_datetime(n.loc[mask_num], unit="D", origin="1899-12-30", errors="coerce")
        out.loc[mask_num] = parsed

    # Strings + generic parse
    rest_idx = (~mask_num) | mask_num.isna()
    if rest_idx.any():
        rest = col[rest_idx]
        rest = rest.astype(str).str.strip().replace({"": None, "nan": None})
        for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%m/%d/%y", "%d/%m/%Y", "%d-%b-%Y"):
            parsed = pd.to_datetime(rest, format=fmt, errors="coerce")
            out = out.combine_first(parsed)
        with warnings.catch_warnings():
            warnings.simplefilter("ignore", category=UserWarning)
            parsed_generic = pd.to_datetime(rest, errors="coerce")
        out = out.combine_first(parsed_generic)

    # Sanity filter
    lo = pd.Timestamp(year=2000, month=1, day=1, tz=None)
    hi = pd.Timestamp(datetime.utcnow() + timedelta(days=31))
    out = out.where((out >= lo) & (out <= hi))
    return pd.to_datetime(out, errors="coerce")

# =========================
#   Period helpers
# =========================
def _coerce_period_end(dN_from_data: datetime, query_val: str | None, dates_for_scan: list[datetime] | None = None, year_qs: str | None = None) -> datetime:
    if year_qs:
        try:
            y = int(str(year_qs).strip())
            if dates_for_scan:
                cand = [d for d in dates_for_scan if d.year == y]
                if cand:
                    return max(cand)
        except Exception:
            pass

    if query_val:
        s = query_val.strip()
        if len(s) == 7 and s[4] == "-":  # YYYY-MM
            try:
                y, m = int(s[:4]), int(s[5:7])
                if dates_for_scan:
                    cand = [d for d in dates_for_scan if d.year == y and d.month == m]
                    if cand:
                        return max(cand)
            except Exception:
                pass
        if len(s) == 4 and s.isdigit():  # YYYY
            try:
                y = int(s)
                if dates_for_scan:
                    cand = [d for d in dates_for_scan if d.year == y]
                    if cand:
                        return max(cand)
            except Exception:
                pass
        try:
            return datetime.fromisoformat(s)  # YYYY-MM-DD
        except Exception:
            pass

    return dN_from_data

def _quarter_start(dt: datetime) -> datetime:
    q = (dt.month - 1) // 3
    m = q * 3 + 1
    return datetime(dt.year, m, 1)

def _ytd_start(dt: datetime) -> datetime:
    return datetime(dt.year, 1, 1)

def _bounds_for_basis(dates: list[datetime], basis: str, period_end_qs: str | None, year_qs: str | None = None):
    if not dates:
        return (None, None)
    d0 = min(dates)
    dN = max(dates)
    end = _coerce_period_end(dN, period_end_qs, dates_for_scan=dates, year_qs=year_qs)
    basis = (basis or "inception").lower()
    if basis == "latest":
        start = end
    elif basis == "inception":
        start = d0
    elif basis == "ytd":
        start = _ytd_start(end)
    elif basis == "quarter":
        start = _quarter_start(end)
    elif basis in ("month", "day"):
        start = end
    else:
        start = d0
    if end > dN:
        end = dN
    return (start, end)

def _span_dict(start_dt, end_dt):
    if not start_dt or not end_dt:
        return None
    start = pd.to_datetime(start_dt)
    end = pd.to_datetime(end_dt)
    days = int((end - start).days)
    years = round(days / 365.25, 6)
    return {"start_date": start.isoformat(), "end_date": end.isoformat(), "days": days, "years": years}

def _irr_from_span(initial_value: float, current_value: float, start_dt, end_dt):
    try:
        if not initial_value or initial_value <= 0 or current_value is None:
            return None
        if not start_dt or not end_dt:
            return None
        years = max((pd.to_datetime(end_dt) - pd.to_datetime(start_dt)).days / 365.25, 0.0)
        if years <= 0:
            return None
        return (pow(float(current_value) / float(initial_value), 1.0 / years) - 1.0) * 100.0
    except Exception:
        return None

# =========================
#   DB-first helpers
# =========================
def _db_overview_latest(sheet: str):
    """
    Use snapshots table if available (InvestorBalance/WorkbookSnapshot), else None.
    """
    if not WorkbookSnapshot or not InvestorBalance:
        return None
    try:
        from sqlalchemy import desc
        snap = (WorkbookSnapshot.query.filter_by(sheet=sheet)
                .order_by(desc(WorkbookSnapshot.as_of), desc(WorkbookSnapshot.id))
                .first())
        if not snap:
            return None
        rows = InvestorBalance.query.filter_by(snapshot_id=snap.id).all()
        if not rows:
            return None
        current_total = sum((r.current_value or 0.0) for r in rows)
        initial_total = sum((r.initial_value or 0.0) for r in rows)
        moic = (current_total / initial_total) if initial_total else None
        roi_pct = ((current_total - initial_total) / initial_total * 100.0) if initial_total else None
        return {
            "source": "db",
            "sheet": sheet,
            "latest_date": snap.as_of.isoformat(),
            "row_count": len(rows),
            "ending_balance_total": current_total,
            "current_value": current_total,
            "initial_value": initial_total,
            "committed": initial_total,
            "moic": moic,
            "roi_pct": roi_pct,
            "irr_pct": None,
            "time_span": None,
        }
    except Exception:
        return None

def _resolve_investor_name_for_user(user: User | None) -> str | None:
    if not user:
        return None
    try:
        inv = Investor.query.filter_by(account_user_id=user.id).order_by(Investor.updated_at.desc()).first()
        if inv and inv.name:
            return inv.name.strip()
    except Exception:
        pass
    try:
        if getattr(user, "email", None):
            inv = Investor.query.filter_by(email=user.email).order_by(Investor.updated_at.desc()).first()
            if inv and inv.name:
                return inv.name.strip()
    except Exception:
        pass
    fn = getattr(user, "first_name", "") or ""
    ln = getattr(user, "last_name", "") or ""
    full = f"{fn} {ln}".strip()
    return full or None

def _db_investor_latest(investor: str, sheet: str):
    if not WorkbookSnapshot or not InvestorBalance:
        return None
    try:
        from sqlalchemy import desc
        snap = (WorkbookSnapshot.query.filter_by(sheet=sheet)
                .order_by(desc(WorkbookSnapshot.as_of), desc(WorkbookSnapshot.id))
                .first())
        if not snap:
            return None
        row = (InvestorBalance.query
               .filter_by(snapshot_id=snap.id)
               .filter(InvestorBalance.investor.ilike(f"%{investor}%"))
               .first())
        if not row:
            return None
        start_dt = getattr(row, "initial_date", None)
        end_dt   = getattr(row, "current_date", None)
        irr = getattr(row, "irr_pct", None)
        if irr is None:
            irr = _irr_from_span(row.initial_value, row.current_value, start_dt, end_dt)
        return {
            "source": "db",
            "sheet": sheet,
            "latest_date": (end_dt.isoformat() if end_dt else snap.as_of.isoformat()),
            "investor": investor,
            "initial_value": row.initial_value,
            "current_value": row.current_value,
            "moic": row.moic,
            "roi_pct": row.roi_pct,
            "irr_pct": irr,
            "time_span": _span_dict(start_dt, end_dt),
            "start_date": (start_dt.isoformat() if start_dt else None),
            "end_date": (end_dt.isoformat() if end_dt else None),
        }
    except Exception:
        return None

# =========================
#   SharePoint LIVE compute
# =========================
def _compute_from_values(values: list[list], sheet_name: str, basis: str = "inception",
                         period_end_qs: str | None = None, year_qs: str | None = None):
    """
    Robust month/YTD/inception computation from Excel usedRange values.
    If no per-row date column is present, detect a single control date in header and use it.
    """
    if not values or len(values) < 2:
        raise RuntimeError("No data returned from Excel API")

    # 1) Header row
    header_row_idx = 0
    for i in range(min(50, len(values))):
        row = [str(c).strip().lower() if c is not None else "" for c in values[i]]
        if any(("ending" in c and "balance" in c) or ("current" in c and "value" in c) for c in row):
            header_row_idx = i
            break

    headers_raw = [("" if h is None else str(h).strip()) for h in values[header_row_idx]]
    records = values[header_row_idx + 1:]
    if not records:
        raise RuntimeError("No rows beneath header")

    headers_norm = [_normhdr(h) for h in headers_raw]
    df = pd.DataFrame(records, columns=headers_raw)

    def _find_last_index(cands_norm: tuple[str, ...]):
        idx = None
        for i, h in enumerate(headers_norm):
            if any(c in h for c in cands_norm):
                idx = i
        return idx

    end_i  = _find_last_index(_ENDING_ALIASES_N)
    beg_i  = _find_last_index(_BEGIN_ALIASES_N)
    date_i = _find_last_index(_DATE_ALIASES_N)
    if end_i is None:
        raise RuntimeError("Could not find an 'Ending/Closing Balance' column")

    # numeric coercion
    def _to_num(x):
        if x is None: return math.nan
        if isinstance(x, (int, float)): return float(x)
        s = str(x).strip().replace(",", "").replace("$", "")
        if s.startswith("(") and s.endswith(")"): s = "-" + s[1:-1]
        try: return float(s)
        except Exception: return math.nan

    df.iloc[:, end_i] = pd.to_numeric(df.iloc[:, end_i].map(_to_num), errors="coerce")
    if beg_i is not None:
        df.iloc[:, beg_i] = pd.to_numeric(df.iloc[:, beg_i].map(_to_num), errors="coerce")

    # 2) Dates: prefer a column; else detect a single control-date in the header
    control_date: datetime | None = None

    if date_i is None:
        def _auto_date_index(frame, exclude_idx):
            best_i, best_hits, n = None, -1, len(frame)
            for i in range(frame.shape[1]):
                if i in exclude_idx:
                    continue
                s = _coerce_dates_series(frame.iloc[:, i])
                hits = int(s.notna().sum())
                ratio = hits / max(1, n)
                if hits >= 4 and ratio >= 0.35:
                    if hits > best_hits:
                        best_i, best_hits = i, hits
            return best_i

        date_i = _auto_date_index(df, exclude_idx={end_i} | ({beg_i} if beg_i is not None else set()))

        if date_i is None:
            MAX_SCAN_R = min(40, len(values))
            MAX_SCAN_C = min(80, max(len(r) for r in values[:MAX_SCAN_R]))
            label_hits = {"endingdate", "asofdate", "asof", "period", "month", "date"}

            for r in range(0, MAX_SCAN_R):
                row = values[r]
                for c in range(0, min(MAX_SCAN_C, len(row))):
                    v = row[c]
                    d = _parse_excel_date(v)
                    if not _is_sane_date(d):
                        continue
                    # nearby label?
                    neigh = []
                    for dc in (-2, -1, 0, 1, 2):
                        cc = c + dc
                        if 0 <= cc < len(row):
                            neigh.append(_normhdr(str(row[cc]) if row[cc] is not None else ""))
                    txt = "".join(neigh)
                    if any(lbl in txt for lbl in label_hits) or header_row_idx > r:
                        control_date = d
                        break
                if control_date:
                    break

            if control_date is None:
                found = []
                for r in range(0, MAX_SCAN_R):
                    for v in values[r]:
                        d = _parse_excel_date(v)
                        if _is_sane_date(d):
                            found.append(d)
                if found:
                    control_date = max(found)

    if date_i is not None:
        dates_series = _coerce_dates_series(df.iloc[:, date_i])
        df["__dt"] = dates_series
    else:
        if control_date is None:
            raise RuntimeError("No date column or control date found in the sheet header.")
        df["__dt"] = pd.Series(pd.to_datetime([control_date] * len(df)), index=df.index)

    # keep numeric/date rows
    numeric_cols = [end_i] + ([beg_i] if beg_i is not None else [])
    work = df[(df.iloc[:, numeric_cols].apply(lambda r: r.notna().any(), axis=1)) & (df["__dt"].notna())].copy()
    if work.empty:
        raise RuntimeError("No numeric rows with valid dates found.")

    work["__end"] = pd.to_numeric(work.iloc[:, end_i], errors="coerce")
    if beg_i is not None:
        work["__beg"] = pd.to_numeric(work.iloc[:, beg_i], errors="coerce")

    gb = work.groupby(work["__dt"].dt.date)
    end_by_date = gb["__end"].sum(min_count=1).fillna(0.0).to_dict()
    beg_by_date = gb["__beg"].sum(min_count=1).fillna(0.0).to_dict() if "__beg" in work.columns else {}

    unique_dates = sorted(set(end_by_date.keys()) | set(beg_by_date.keys()))
    start_dt, end_dt = _bounds_for_basis([datetime.combine(d, datetime.min.time()) for d in unique_dates],
                                         basis, period_end_qs, year_qs)
    latest_iso = end_dt.date().isoformat()

    def latest_in_month(dtdate: date):
        ys, ms = dtdate.year, dtdate.month
        c = [d for d in unique_dates if d.year == ys and d.month == ms]
        return max(c) if c else None

    def sum_beg_in_month(dtdate: date):
        if not beg_by_date:
            return None
        ys, ms = dtdate.year, dtdate.month
        c = [d for d in unique_dates if d.year == ys and d.month == ms and d in beg_by_date]
        return float(sum(beg_by_date.get(d, 0.0) for d in c)) if c else None

    def prev_month_latest(dtdate: date):
        y, m = dtdate.year, dtdate.month
        if m == 1: y, m = y - 1, 12
        else: m -= 1
        c = [d for d in unique_dates if d.year == y and d.month == m]
        return max(c) if c else None

    eff_basis = "month" if basis == "latest" else basis
    if eff_basis == "month":
        ek = latest_in_month(end_dt.date())
        if ek is None:
            raise RuntimeError("No data for selected month.")
        current_value = float(end_by_date.get(ek, 0.0))

        month_beg_total = sum_beg_in_month(end_dt.date())
        if month_beg_total is not None:
            initial_value = month_beg_total
        else:
            pk = prev_month_latest(end_dt.date())
            if pk is None:
                first = min(unique_dates)
                initial_value = float(end_by_date.get(first, 0.0))
            else:
                initial_value = float(end_by_date.get(pk, 0.0))
    else:
        ek = latest_in_month(end_dt.date()) if eff_basis in ("ytd", "quarter") else end_dt.date()
        if ek not in end_by_date:
            cands = [d for d in unique_dates if d <= ek]
            ek = max(cands) if cands else unique_dates[-1]
        current_value = float(end_by_date.get(ek, 0.0))

        if beg_by_date:
            cands = [d for d in sorted(unique_dates) if d >= start_dt.date() and d in beg_by_date]
            if cands:
                initial_value = float(beg_by_date.get(cands[0], 0.0))
            else:
                prev = max([d for d in unique_dates if d < start_dt.date()], default=None)
                initial_value = float(end_by_date.get(prev, 0.0)) if prev else 0.0
        else:
            prev = max([d for d in unique_dates if d < ek], default=None)
            initial_value = float(end_by_date.get(prev, 0.0)) if prev else 0.0

    moic = roi_pct = irr_pct = None
    if initial_value and initial_value != 0:
        moic = round(current_value / initial_value, 4)
        roi_pct = round(((current_value - initial_value) / initial_value) * 100.0, 2)
        irr_pct = _irr_from_span(initial_value, current_value, start_dt, end_dt)

    return {
        "source": "sharepoint-live",
        "sheet": sheet_name,
        "basis": eff_basis,
        "period_end": latest_iso,
        "current_value": current_value,
        "initial_value": initial_value,
        "moic": moic,
        "roi_pct": roi_pct,
        "irr_pct": irr_pct,
        "time_span": _span_dict(start_dt, end_dt),
    }

# =========================
#   Openpyxl overview (uploads)
# =========================
def _fast_overview(xlsm_path: Path, sheet_name: str, basis: str = "inception",
                   period_end_qs: str | None = None, year_qs: str | None = None):
    wb_vals = load_workbook(xlsm_path, data_only=True, read_only=True)
    wb_forms = load_workbook(xlsm_path, data_only=False, read_only=True)
    try:
        # resolve sheet name
        target_name = sheet_name
        if target_name not in wb_vals.sheetnames:
            # simple normalization pass
            want = _normalize_sheet_name(target_name)
            for s in wb_vals.sheetnames:
                if _normalize_sheet_name(s) == want or want in _normalize_sheet_name(s):
                    target_name = s
                    break
        ws_v = wb_vals[target_name]
        ws_f = wb_forms[target_name]

        # Find basic headers
        # Reuse helpers from second version
        asof_row, asof_col, beg_row, beg_cols, end_row, end_cols = _find_header_row_and_cols(ws_v)
        if not end_cols:
            raise RuntimeError("Could not find an 'Ending/Closing Balance' column.")
        target_end_col = end_cols[-1]
        target_beg_col = beg_cols[-1] if beg_cols else None

        date_to_end, date_to_beg = {}, {}
        if asof_col:
            for r in range((asof_row or 1) + 1, ws_v.max_row + 1):
                d = _parse_excel_date(ws_v.cell(row=r, column=asof_col).value)
                if not _is_sane_date(d):
                    continue
                dkey = d.date()
                ev = _to_float(ws_v.cell(row=r, column=target_end_col).value)
                if not math.isnan(ev):
                    date_to_end[dkey] = date_to_end.get(dkey, 0.0) + float(ev)
                if target_beg_col:
                    bv = _to_float(ws_v.cell(row=r, column=target_beg_col).value)
                    if not math.isnan(bv):
                        date_to_beg[dkey] = date_to_beg.get(dkey, 0.0) + float(bv)

        if not date_to_end:
            # Fallback: scan TOTAL row for current value & committed
            # This mirrors the simpler approach from the other file
            max_scan_rows = min(300, ws_v.max_row)
            max_scan_cols = min(300, ws_v.max_column)
            # find "Ending Balance" header row
            ending_row = None; ending_cols = []
            committed_row = None; committed_cols = []
            for r in range(1, max_scan_rows + 1):
                texts = _row_texts(ws_v, r, max_scan_cols)
                lower = [t.lower() for t in texts]
                eb_cols = [i + 1 for i, t in enumerate(lower) if "ending balance" in t]
                cm_cols = [i + 1 for i, t in enumerate(lower) if re.search(r"\bcommitted\b|\bcommitment\b", t)]
                if eb_cols and ending_row is None:
                    ending_row = r; ending_cols = eb_cols
                if cm_cols and committed_row is None:
                    committed_row = r; committed_cols = cm_cols
                if ending_row and committed_row:
                    break

            if not ending_row or not ending_cols:
                raise RuntimeError("No dates nor basic headers found.")
            target_ending_col = ending_cols[-1]

            total_row = _scan_total_row_uploads(ws_v, start_row=ending_row + 1, label_cols=12)
            if total_row:
                cv = _to_float(ws_v.cell(row=total_row, column=target_ending_col).value)
                current_value = float(0 if math.isnan(cv) else cv)
            else:
                current_value = _sum_contiguous_block_uploads(ws_v, start_row=ending_row + 1, col=target_ending_col)

            initial_value = None
            if committed_row and committed_cols:
                cm_total_row = _scan_total_row_uploads(ws_v, start_row=committed_row + 1, label_cols=12)
                if cm_total_row:
                    acc = _to_float(ws_v.cell(row=cm_total_row, column=committed_cols[-1]).value)
                    if not math.isnan(acc):
                        initial_value = float(acc)
            if initial_value is None:
                # heuristic: best committed figure
                best = math.nan
                if committed_row and committed_cols:
                    for rr in range(committed_row + 1, min(ws_v.max_row, committed_row + 1000) + 1):
                        v = _to_float(ws_v.cell(row=rr, column=committed_cols[-1]).value)
                        if not math.isnan(v) and (math.isnan(best) or v > best):
                            best = v
                    if not math.isnan(best):
                        initial_value = float(best)

            moic = roi_pct = None
            if initial_value:
                moic = float(current_value) / float(initial_value)
                roi_pct = (moic - 1.0) * 100.0

            return {
                "initial_value": float(initial_value) if initial_value is not None else None,
                "current_value": float(current_value),
                "moic": moic,
                "roi_pct": roi_pct,
                "irr_pct": None,
                "time_span": None,
                "header": {"latest_date": None},
                "excel": {"ending": {"col_letter": get_column_letter(target_ending_col)}},
                "file": str(xlsm_path),
                "sheet": target_name,
                "basis": basis,
                "period_end": None,
            }

        # With date_to_* maps available, reuse basis computation
        all_dates = sorted(set(date_to_end.keys()) | set(date_to_beg.keys()))
        all_dt = [datetime.combine(d, datetime.min.time()) for d in all_dates]
        start_dt, end_dt = _bounds_for_basis(all_dt, basis, period_end_qs, year_qs)
        latest_iso = end_dt.date().isoformat()

        def _latest_in_month(dtdate):
            ys, ms = dtdate.year, dtdate.month
            cands = [d for d in all_dates if d.year == ys and d.month == ms]
            return max(cands) if cands else None

        def _sum_beg_in_month(dtdate):
            ys, ms = dtdate.year, dtdate.month
            cands = [d for d in all_dates if d.year == ys and d.month == ms and d in date_to_beg]
            return float(sum(date_to_beg.get(d, 0.0) for d in cands)) if cands else None

        def _prev_month_latest(dtdate):
            y, m = dtdate.year, dtdate.month
            if m == 1: y, m = y - 1, 12
            else: m -= 1
            cands = [d for d in all_dates if d.year == y and d.month == m]
            return max(cands) if cands else None

        eff_basis = "month" if basis == "latest" else basis
        if eff_basis == "month":
            ek = _latest_in_month(end_dt.date())
            if ek is None:
                raise RuntimeError("No data for selected month.")
            current_value = float(date_to_end.get(ek, 0.0))
            month_beg_total = _sum_beg_in_month(end_dt.date()) if target_beg_col else None
            if month_beg_total is not None:
                initial_value = month_beg_total
            else:
                prev_key = _prev_month_latest(end_dt.date())
                if prev_key is None:
                    first = min(all_dates)
                    initial_value = float(date_to_end.get(first, 0.0))
                else:
                    initial_value = float(date_to_end.get(prev_key, 0.0))
        else:
            ek = _latest_in_month(end_dt.date()) if eff_basis in ("ytd", "quarter") else end_dt.date()
            if ek not in date_to_end:
                cands = [d for d in all_dates if d <= ek]
                ek = max(cands) if cands else all_dates[-1]
            current_value = float(date_to_end.get(ek, 0.0))
            if target_beg_col:
                start_key = min([d for d in all_dates if d >= start_dt.date() and d in date_to_beg], default=None)
                if start_key:
                    initial_value = float(date_to_beg.get(start_key, 0.0))
                else:
                    prev_key = max([d for d in all_dates if d < start_dt.date()], default=None)
                    initial_value = float(date_to_end.get(prev_key, 0.0)) if prev_key else 0.0
            else:
                prev = max([d for d in all_dates if d < ek], default=None)
                initial_value = float(date_to_end.get(prev, 0.0)) if prev else 0.0

        moic = roi_pct = irr_pct = None
        if initial_value and initial_value != 0:
            moic = float(current_value) / float(initial_value)
            roi_pct = (moic - 1.0) * 100.0
            irr_pct = _irr_from_span(initial_value, current_value, start_dt, end_dt)

        return {
            "basis": eff_basis,
            "period_end": latest_iso,
            "initial_value": float(initial_value),
            "current_value": float(current_value),
            "moic": moic,
            "roi_pct": roi_pct,
            "irr_pct": irr_pct,
            "time_span": _span_dict(start_dt, end_dt),
            "file": str(xlsm_path),
            "sheet": target_name,
        }
    finally:
        try: wb_vals.close()
        except Exception: pass
        try: wb_forms.close()
        except Exception: pass

def _overview_cached(xlsm_path: Path, sheet: str, basis: str, period_end_iso: str | None, year_qs: str | None):
    key = (str(xlsm_path), sheet, basis.lower(), period_end_iso or "", year_qs or "")
    mtime = os.path.getmtime(xlsm_path)
    entry = _OVERVIEW_CACHE.get(key)
    if entry and entry["mtime"] == mtime:
        return entry["data"]
    data = _fast_overview(xlsm_path, sheet, basis=basis, period_end_qs=period_end_iso, year_qs=year_qs)
    _OVERVIEW_CACHE[key] = {"mtime": mtime, "ts": time.time(), "data": data}
    return data

# =========================
#   Persist-on-read (DB)
# =========================
def _upsert_period_metric(sheet_name: str, data: dict):
    """
    Persist the latest computed month totals whenever overview runs.
    """
    try:
        pe = str(data.get("period_end") or "").strip()
        if not pe:
            return
        # Accept YYYY-MM or YYYY-MM-DD
        if len(pe) == 7 and pe[4] == "-":
            y, m = int(pe[:4]), int(pe[5:7])
            if m == 12: d = 31
            else:
                from calendar import monthrange
                d = monthrange(y, m)[1]
            as_of = date(y, m, d)
        else:
            as_of = datetime.fromisoformat(pe).date()

        init_v = data.get("initial_value")
        end_v  = data.get("current_value")
        if init_v is None or end_v is None:
            return

        stmt = sqlite_insert(PortfolioPeriodMetric).values(
            sheet=sheet_name,
            as_of_date=as_of,
            beginning_balance=float(init_v),
            ending_balance=float(end_v),
            unrealized_gain_loss=data.get("unrealized_gain_loss"),
            realized_gain_loss=data.get("realized_gain_loss"),
            management_fees=data.get("management_fees"),
            source=data.get("source") or "overview",
        )
        stmt = stmt.on_conflict_do_update(
            index_elements=["sheet", "as_of_date"],
            set_={
                "beginning_balance": stmt.excluded.beginning_balance,
                "ending_balance": stmt.excluded.ending_balance,
                "unrealized_gain_loss": stmt.excluded.unrealized_gain_loss,
                "realized_gain_loss": stmt.excluded.realized_gain_loss,
                "management_fees": stmt.excluded.management_fees,
                "updated_at": datetime.utcnow(),
                "source": stmt.excluded.source,
            },
        )
        db.session.execute(stmt)
        db.session.commit()
    except Exception as e:
        current_app.logger.exception("overview upsert failed: %s", e)
        db.session.rollback()

# =========================
#   Utility: current user
# =========================
def _current_user_id():
    return request.headers.get("X-User-Id") or _session.get("user_id")

class SimpleUser:
    def __init__(self, id=None, email=None, first_name=None, last_name=None, name=None):
        self.id = id
        self.email = email
        self.first_name = first_name
        self.last_name = last_name
        if name and not (first_name or last_name):
            parts = str(name).strip().split(" ")
            if len(parts) >= 2:
                self.first_name = parts[0]
                self.last_name = " ".join(parts[1:])
            else:
                self.first_name = name
                self.last_name = ""

def _current_user():
    uid = _current_user_id()
    if uid:
        try:
            u = User.query.filter_by(id=uid).first()
            if u:
                return u
        except Exception:
            pass
    email = request.headers.get("X-User-Email")
    name  = request.headers.get("X-User-Name")
    if email or name:
        return SimpleUser(id=None, email=email, name=name)
    return None

def _latest_sharepoint_connection_for_user():
    uid = _current_user_id()
    if not uid:
        return None
    try:
        return (
            SharePointConnection.query
            .filter_by(user_id=uid)
            .order_by(SharePointConnection.id.desc())
            .first()
        )
    except Exception:
        return None

def _latest_sharepoint_connection_for_user_or_shared():
    own = _latest_sharepoint_connection_for_user()
    if own:
        return own
    try:
        if hasattr(SharePointConnection, "is_shared"):
            shared = (
                SharePointConnection.query
                .filter_by(is_shared=True)
                .order_by(SharePointConnection.id.desc())
                .first()
            )
            if shared:
                return shared
    except Exception:
        pass
    return None

# =========================
#   API: DB month read
# =========================
@metrics_bp.get("/overview/db")
def overview_from_db():
    """
    Read a month snapshot from portfolio_period_metrics.

    Modes:
      • If period_end=YYYY-MM (or YYYY-MM-DD) is provided: return that month.
      • If period_end is omitted: span mode → initial=current of earliest row; current=ending of latest row.
    """
    from calendar import monthrange

    sheet = request.args.get("sheet", "bCAS (Q4 Adj)")
    pe = (request.args.get("period_end") or "").strip()

    # span mode
    if not pe:
        rows = (
            PortfolioPeriodMetric.query
            .filter_by(sheet=sheet)
            .order_by(PortfolioPeriodMetric.as_of_date.asc())
            .all()
        )
        if not rows:
            return jsonify(error="no data for requested sheet"), 404

        first, last = rows[0], rows[-1]
        initial_value = float(first.ending_balance or 0.0)
        current_value = float(last.ending_balance or 0.0)
        moic = (current_value / initial_value) if initial_value else None
        roi_pct = ((current_value - initial_value) / initial_value) * 100.0 if initial_value else None

        return jsonify({
            "source": "db",
            "sheet": sheet,
            "basis": "span",
            "period_end": f"{last.as_of_date.year}-{str(last.as_of_date.month).zfill(2)}",
            "initial_value": initial_value,
            "current_value": current_value,
            "moic": moic,
            "roi_pct": roi_pct,
        })

    # month mode
    try:
        if len(pe) == 7 and pe[4] == "-":
            y, m = int(pe[:4]), int(pe[5:7])
            day = monthrange(y, m)[1]
            as_of = date(y, m, day)
        else:
            as_of = datetime.fromisoformat(pe).date()
            y, m = as_of.year, as_of.month
            as_of = date(y, m, monthrange(y, m)[1])
    except Exception:
        return jsonify(error="invalid period_end"), 400

    row = (
        PortfolioPeriodMetric.query
        .filter_by(sheet=sheet, as_of_date=as_of)
        .first()
    )
    if not row:
        return jsonify(error="no data for requested month"), 404

    initial_value = row.beginning_balance
    if initial_value is None:
        prev_y, prev_m = (as_of.year - 1, 12) if as_of.month == 1 else (as_of.year, as_of.month - 1)
        prev_day = monthrange(prev_y, prev_m)[1]
        prev_asof = date(prev_y, prev_m, prev_day)
        prev = (
            PortfolioPeriodMetric.query
            .filter_by(sheet=sheet, as_of_date=prev_asof)
            .first()
        )
        initial_value = prev.ending_balance if prev and prev.ending_balance is not None else 0.0

    current_value = row.ending_balance or 0.0
    moic = (current_value / initial_value) if initial_value else None
    roi_pct = ((current_value - initial_value) / initial_value) * 100.0 if initial_value else None

    return jsonify({
        "source": "db",
        "sheet": sheet,
        "basis": "month",
        "period_end": f"{as_of.year}-{str(as_of.month).zfill(2)}",
        "initial_value": float(initial_value) if initial_value is not None else None,
        "current_value": float(current_value) if current_value is not None else None,
        "moic": moic,
        "roi_pct": roi_pct,
    })

# =========================
#   API: Overview (uploads → SharePoint), with persist-on-read
# =========================
@metrics_bp.get("/overview")
def overview():
    """
    Overview resolver (uploads first, else SharePoint LIVE)
    Query:
      ?sheet=...
      ?file=...                 (uploads)
      ?conn_id=... | ?url=...   (SharePoint live)
      ?basis=inception|ytd|quarter|month|day|latest
      ?period_end=YYYY | YYYY-MM | YYYY-MM-DD
      ?year=YYYY
    """
    sheet = request.args.get("sheet", "bCAS (Q4 Adj)")
    filename = request.args.get("file")
    conn_id = request.args.get("conn_id")
    share_url = (request.args.get("url") or "").strip()
    basis = (request.args.get("basis") or "inception").lower()
    period_end_iso = request.args.get("period_end")
    year_qs = request.args.get("year")

    if basis not in {"inception","ytd","quarter","month","day","latest"}:
        basis = "inception"

    # 1) Local uploads
    try:
        uploads = _uploads_dir()
        xlsm_path = _find_best_xlsm(uploads, filename)
        data = _overview_cached(xlsm_path, sheet, basis=basis, period_end_iso=period_end_iso, year_qs=year_qs)
        _upsert_period_metric(sheet, data)
        resp = jsonify(data); resp.headers["Cache-Control"] = "no-store"
        return resp
    except FileNotFoundError:
        pass
    except Exception:
        logging.exception("[metrics] uploads path failed; trying SharePoint")

    # 2) SharePoint LIVE
    try:
        bearer = _resolve_bearer()
        if not bearer:
            up_dir = str(_uploads_dir())
            return _unauthorized_response_no_bearer("overview (aggregate)", searched_uploads=up_dir)

        # Resolve drive/item
        if conn_id:
            uid = _current_user_id()
            if not uid: return jsonify(error="Not authenticated"), 401
            conn = SharePointConnection.query.filter_by(id=conn_id, user_id=uid).first()
            if not conn: return jsonify(error="SharePoint connection not found"), 404
            drive_id, item_id = conn.drive_id, conn.item_id
        elif share_url:
            drive_id, item_id = open_excel_by_share_url(share_url, bearer)
        else:
            conn = _latest_sharepoint_connection_for_user_or_shared()
            if not conn:
                return jsonify(error="No uploaded workbook and no saved SharePoint connection."), 400
            drive_id, item_id = conn.drive_id, conn.item_id

        sid = create_session(drive_id, item_id, bearer, persist=False)
        try:
            # Optional: fast control-cell path for month/latest when workbook has control sheet cells configured
            if (
                basis in ("month","latest") and
                (period_end_iso or basis == "latest") and
                current_app.config.get("METRICS_CTRL_SHEET") and
                current_app.config.get("METRICS_CTRL_YEAR") and
                current_app.config.get("METRICS_CTRL_MONTH") and
                current_app.config.get("METRICS_CELL_END") and
                write_range_value is not None
            ):
                ctrl_sheet = current_app.config["METRICS_CTRL_SHEET"]
                year_addr  = current_app.config.get("METRICS_CTRL_YEAR")
                month_addr = current_app.config.get("METRICS_CTRL_MONTH")
                begin_addr = current_app.config.get("METRICS_CELL_BEGIN")  # optional
                end_addr   = current_app.config["METRICS_CELL_END"]

                try:
                    if basis == "latest" and not period_end_iso:
                        raise RuntimeError("force generic scan for latest to detect max date")

                    y, m = period_end_iso.split("-")[:2]
                    write_range_value(drive_id, item_id, ctrl_sheet, year_addr, int(y), bearer, sid)
                    write_range_value(drive_id, item_id, ctrl_sheet, month_addr, int(m), bearer, sid)

                    def _read_one(addr):
                        payload = sp_read_range(drive_id, item_id, ctrl_sheet, addr, bearer)
                        vals = (payload or {}).get("values") or [[]]
                        v = vals[0][0] if vals and vals[0] else None
                        s = str(v).replace(",", "").replace("$", "").strip() if v is not None else ""
                        if s.startswith("(") and s.endswith(")"): s = "-" + s[1:-1]
                        try: return float(s)
                        except Exception: return 0.0

                    current_value = _read_one(end_addr)
                    begin_val = _read_one(begin_addr) if begin_addr else None
                    if begin_val is not None and begin_val != 0.0:
                        initial_value = begin_val
                    else:
                        prev_y, prev_m = int(y), int(m)
                        if prev_m == 1:
                            prev_y -= 1; prev_m = 12
                        else:
                            prev_m -= 1
                        write_range_value(drive_id, item_id, ctrl_sheet, year_addr, int(prev_y), bearer, sid)
                        write_range_value(drive_id, item_id, ctrl_sheet, month_addr, int(prev_m), bearer, sid)
                        time.sleep(0.3)
                        initial_value = _read_one(end_addr)
                        write_range_value(drive_id, item_id, ctrl_sheet, year_addr, int(y), bearer, sid)
                        write_range_value(drive_id, item_id, ctrl_sheet, month_addr, int(m), bearer, sid)

                    moic = (current_value / initial_value) if initial_value else None
                    roi_pct = ((current_value - initial_value) / initial_value) * 100.0 if initial_value else None

                    data = {
                        "source": "sharepoint-live",
                        "sheet": sheet,
                        "basis": "month",
                        "period_end": f"{y}-{m}",
                        "initial_value": initial_value,
                        "current_value": current_value,
                        "moic": moic,
                        "roi_pct": roi_pct,
                    }
                    _upsert_period_metric(sheet, data)
                    resp = jsonify(data); resp.headers["Cache-Control"] = "no-store"
                    return resp
                except Exception:
                    pass  # fall through to generic usedRange scan

            # Generic usedRange scan
            values = None; last_err = None; tried = []
            for sn in _sheet_candidates(sheet) or [sheet]:
                tried.append(sn)
                try:
                    values = used_range_values(drive_id, item_id, sn, bearer, sid)
                    sheet = sn; break
                except Exception as e:
                    last_err = e
            if values is None:
                resolved, names = _resolve_remote_sheet_name(drive_id, item_id, sheet, bearer)
                if resolved:
                    try:
                        values = used_range_values(drive_id, item_id, resolved, bearer, sid)
                        sheet = resolved
                    except Exception as e2:
                        last_err = e2
                if values is None:
                    return jsonify(error=(f"Worksheet not found (tried: {tried}). Available sheets: {names or 'unknown'}. Last error: {last_err}")), 400
        finally:
            close_session(drive_id, item_id, bearer, sid)

        data = _compute_from_values(values, sheet, basis=basis, period_end_qs=period_end_iso, year_qs=year_qs)
        _upsert_period_metric(sheet, data)
        resp = jsonify(data); resp.headers["Cache-Control"] = "no-store"
        return resp
    except Exception as e2:
        logging.exception("[metrics] /overview SharePoint fallback failed")
        return jsonify({"error": str(e2)}), 400

# =========================
#   API: Investor overview
# =========================
@metrics_bp.get("/investor-overview")
def investor_overview():
    """
    Investor dashboard (period-aware, per-investor).

    Default (no range): initial = first period's Ending Balance,
                        current = last period's Ending Balance.

    With range (?from, ?to): use InvestorPeriodBalance rows within the
    window; if the first row has no beginning, fall back to the nearest
    prior period's ending for that investor.

    Also returns join_date = the earliest period_date for this investor.
    """
    from calendar import monthrange

    def _month_end(d: date) -> date:
        return date(d.year, d.month, monthrange(d.year, d.month)[1])

    def _parse_monthish(s: str | None):
        if not s:
            return None
        s = s.strip()
        try:
            if len(s) == 7 and s[4] == "-":  # YYYY-MM
                y, m = int(s[:4]), int(s[5:7])
                return date(y, m, monthrange(y, m)[1])
            dt = datetime.fromisoformat(s).date()  # YYYY-MM-DD
            return _month_end(dt)
        except Exception:
            return None

    try:
        me = _current_user()
        resolved_name = _resolve_investor_name_for_user(me)
        investor = (request.args.get("investor") or "").strip() or resolved_name or "Unknown Investor"
        sheet = (request.args.get("sheet") or "bCAS (Q4 Adj)").strip()

        from_qs = _parse_monthish(request.args.get("from"))
        to_qs   = _parse_monthish(request.args.get("to"))

        # ── compute earliest (join) date ONCE
        first_any = (InvestorPeriodBalance.query
                     .filter(InvestorPeriodBalance.investor == investor)
                     .order_by(InvestorPeriodBalance.period_date.asc())
                     .first())
        join_date_iso = first_any.period_date.isoformat() if first_any else None

        def _kpis_from_rows(rows: list[InvestorPeriodBalance], prev_row: InvestorPeriodBalance | None):
            first = rows[0]
            last  = rows[-1]

            init_val = first.beginning_balance
            if init_val is None:
                if prev_row and prev_row.ending_balance is not None:
                    init_val = float(prev_row.ending_balance)
                else:
                    init_val = float(first.ending_balance or 0.0)

            curr_val = float(last.ending_balance or 0.0)

            moic = (curr_val / init_val) if init_val else None
            roi_pct = ((curr_val - init_val) / init_val) * 100.0 if init_val else None
            irr_pct = _irr_from_span(init_val, curr_val, rows[0].period_date, rows[-1].period_date)

            return jsonify({
                "source": "db",
                "sheet": sheet,
                "investor": investor,
                "initial_value": float(init_val),
                "current_value": float(curr_val),
                "moic": moic,
                "roi_pct": roi_pct,
                "irr_pct": irr_pct,
                "time_span": _span_dict(rows[0].period_date, rows[-1].period_date),
                "current_date": rows[-1].period_date.isoformat(),
                "join_date": join_date_iso,  # ← NEW
            })

        # ── Period-selected path
        if from_qs and to_qs and from_qs <= to_qs:
            rows = (InvestorPeriodBalance.query
                    .filter(InvestorPeriodBalance.investor == investor)
                    .filter(InvestorPeriodBalance.period_date >= from_qs)
                    .filter(InvestorPeriodBalance.period_date <= to_qs)
                    .order_by(InvestorPeriodBalance.period_date.asc())
                    .all())
            if rows:
                prev = (InvestorPeriodBalance.query
                        .filter(InvestorPeriodBalance.investor == investor)
                        .filter(InvestorPeriodBalance.period_date < rows[0].period_date)
                        .order_by(InvestorPeriodBalance.period_date.desc())
                        .first())
                return _kpis_from_rows(rows, prev)

        # ── Default path (no range): span ALL investor periods
        all_rows = (InvestorPeriodBalance.query
                    .filter(InvestorPeriodBalance.investor == investor)
                    .order_by(InvestorPeriodBalance.period_date.asc())
                    .all())
        if all_rows:
            first_row, last_row = all_rows[0], all_rows[-1]
            initial_value = float(first_row.ending_balance or 0.0)
            current_value = float(last_row.ending_balance or 0.0)
            moic = (current_value / initial_value) if initial_value else None
            roi_pct = ((current_value - initial_value) / initial_value) * 100.0 if initial_value else None
            irr_pct = _irr_from_span(initial_value, current_value, first_row.period_date, last_row.period_date)

            return jsonify({
                "source": "db",
                "sheet": sheet,
                "investor": investor,
                "initial_value": initial_value,
                "current_value": current_value,
                "moic": moic,
                "roi_pct": roi_pct,
                "irr_pct": irr_pct,
                "time_span": _span_dict(first_row.period_date, last_row.period_date),
                "current_date": last_row.period_date.isoformat(),
                "join_date": join_date_iso,  # ← NEW
            })

        # ── Fallbacks
        try:
            db_data = _db_investor_latest(investor, sheet)
            if db_data:
                if "start_date" in db_data and db_data["start_date"]:
                    db_data["join_date"] = db_data["start_date"]
                else:
                    db_data["join_date"] = join_date_iso
                db_data["investor"] = investor
                return jsonify(db_data)
        except Exception:
            logging.info("[investor-overview] investor-period DB not available; trying file/Graph")

        return jsonify(error="No investor data available"), 404

    except Exception as e:
        logging.exception("investor-overview failed")
        return jsonify(error=str(e)), 500

# =========================
#   API: Convenience (uploads)
# =========================
@metrics_bp.get("/current-value")
def get_current_value():
    try:
        sheet = request.args.get("sheet", "bCAS (Q4 Adj)")
        basis = (request.args.get("basis") or "inception").lower()
        period_end_iso = request.args.get("period_end")
        year_qs = request.args.get("year")

        # DB-first
        db_data = _db_overview_latest(sheet)
        if db_data and "current_value" in db_data:
            return jsonify({"current_value": db_data["current_value"],
                            "header": {"latest_date": db_data.get("latest_date")}})

        up = _uploads_dir()
        filename = request.args.get("file")
        xlsm_path = _find_best_xlsm(up, filename)
        o = _overview_cached(xlsm_path, sheet, basis=basis, period_end_iso=period_end_iso, year_qs=year_qs)
        _upsert_period_metric(sheet, o)
        return jsonify({"current_value": o["current_value"], "header": o.get("header")})
    except Exception as e:
        logging.exception("[metrics] current-value failed")
        return jsonify({"error": str(e)}), 400

@metrics_bp.get("/initial-value")
def get_initial_value():
    try:
        sheet = request.args.get("sheet", "bCAS (Q4 Adj)")
        basis = (request.args.get("basis") or "inception").lower()
        period_end_iso = request.args.get("period_end")
        year_qs = request.args.get("year")

        # DB-first
        db_data = _db_overview_latest(sheet)
        if db_data and "initial_value" in db_data:
            return jsonify({
                "initial_value": db_data["initial_value"],
                "file": None,
                "excel": {},
            })

        up = _uploads_dir()
        filename = request.args.get("file")
        xlsm_path = _find_best_xlsm(up, filename)
        o = _overview_cached(xlsm_path, sheet, basis=basis, period_end_iso=period_end_iso, year_qs=year_qs)
        _upsert_period_metric(sheet, o)
        committed = o.get("excel", {}).get("committed", {}) if isinstance(o.get("excel"), dict) else {}
        return jsonify({
            "initial_value": o["initial_value"],
            "raw_value": committed.get("raw_value"),
            "formula": committed.get("formula"),
            "excel": committed,
            "file": o.get("file"),
        })
    except Exception as e:
        logging.exception("[metrics] initial-value failed")
        return jsonify({"error": str(e)}), 400

@metrics_bp.get("/debug/uploads")
def debug_uploads():
    up = _uploads_dir()
    return jsonify({
        "root_path": str(Path(current_app.root_path).resolve()),
        "cwd": str(Path.cwd().resolve()),
        "uploads_dir": str(up),
        "default_workbook_file": _cfg("DEFAULT_WORKBOOK_FILE"),
        "xlsm_files": sorted(f.name for f in up.glob("*.xlsm")),
    })

@metrics_bp.get("/files")
def list_files():
    up = _uploads_dir()
    return jsonify(sorted(f.name for f in up.glob("*.xlsm")))

# =========================
#   API: Ingestion (SharePoint → DB monthly)
# =========================
import regex as _re
def _clean_txt(x: str) -> str:
    s = (x or "").replace("\u00a0", " ").replace("\u2011", "-")
    return _re.sub(r"\s+", " ", s).strip().lower()

_LABEL_ALIASES = {
    "Beginning Balance": [
        r"^begin(ning)? balance$",
        r"^opening (nav|balance)$",
        r"^current period begin(ning)? balance$",
        r"^total begin(ning)? balance$",
        r"^total beginning balance$",
    ],
    "Ending Balance": [
        r"^ending balance$",
        r"^closing balance$",
        r"^current value$",
        r"^total ending balance$",
        r"^total current value$",
    ],
    "Unrealized Gain/Loss": [
        r"^unrealis?ed gain/?loss$",
        r"^unrealized pnl$",
        r"^realis?ed gain/\(loss\)$",
        r"^total unrealis?ed gain/?loss$",
    ],
    "Realized Gain/Loss": [
        r"^realis?ed gain/?loss$",
        r"^realized pnl$",
        r"^total realis?ed gain/?loss$",
    ],
    "Management Fees": [
        r"^management fees?$",
        r"^manag(e|ing) fees?$",
        r"^mgmt fees?$",
        r"^total management fees?$",
    ],
}
_METRIC_KEYS = {
    "Beginning Balance": "beginning",
    "Ending Balance": "ending",
    "Unrealized Gain/Loss": "unrealized",
    "Realized Gain/Loss": "realized",
    "Management Fees": "fees",
}

def _find_label_row(values, canonical_label, search_rows=200):
    patterns = [_re.compile(pat, _re.I) for pat in _LABEL_ALIASES.get(canonical_label, [canonical_label])]
    for r_idx, row in enumerate(values[:search_rows], start=1):
        for cell in row:
            txt = _clean_txt(str(cell))
            if any(p.fullmatch(txt) for p in patterns):
                return r_idx
    return None

def _find_total_row(values, search_rows=500, search_cols=30):
    for r in range(1, min(len(values), search_rows) + 1):
        row = values[r - 1]
        stop_c = min(search_cols, len(row))
        for c in range(1, stop_c + 1):
            v = row[c - 1]
            if isinstance(v, str) and v.strip().lower() == "total":
                return r
    return None

def _looks_like_date(v):
    if isinstance(v, (int, float)) and 20000 < float(v) < 90000:
        return True
    if isinstance(v, (datetime, date)): return True
    try:
        datetime.strptime(str(v).strip(), "%m/%d/%Y"); return True
    except Exception: pass
    try:
        datetime.strptime(str(v).strip(), "%Y-%m-%d"); return True
    except Exception: pass
    return False

def _parse_date_any(v):
    if isinstance(v, datetime): return v.date()
    if isinstance(v, date): return v
    if isinstance(v, (int, float)) and 20000 < float(v) < 90000:
        base = datetime(1899,12,30)
        return (base + timedelta(days=int(v))).date()
    for fmt in ("%m/%d/%Y","%Y-%m-%d","%d/%m/%Y"):
        try: return datetime.strptime(str(v).strip(), fmt).date()
        except Exception: pass
    return None

def _find_header_row_and_date_columns(values, max_scan_rows=120, anchor_row=None):
    candidates = []  # [(row_1b, {col: date})]
    rows = len(values)
    cols = max((len(r) for r in values), default=0)
    for r in range(1, min(rows, max_scan_rows) + 1):
        row = values[r-1]
        local = {}
        for c in range(1, cols+1):
            v = row[c-1] if c-1 < len(row) else None
            if _looks_like_date(v):
                d = _parse_date_any(v)
                if d:
                    local[c] = d
        if len(local) >= 2:
            candidates.append((r, local))

    if not candidates:
        return None, {}

    if anchor_row:
        above = [(r, m) for (r, m) in candidates if r <= anchor_row]
        if above:
            above.sort(key=lambda x: (anchor_row - x[0], -len(x[1])))
            return above[0]
    candidates.sort(key=lambda x: (-len(x[1]), x[0]))
    return candidates[0]

def _metric_for_column(values, header_row_1b: int, col_1b: int, lookback_rows: int = 10):
    compiled = {name: [_re.compile(pat, _re.I) for pat in pats] for name, pats in _LABEL_ALIASES.items()}
    r = header_row_1b
    while r >= 1 and r >= header_row_1b - lookback_rows:
        row = values[r - 1] if r - 1 < len(values) else []
        cell = row[col_1b - 1] if col_1b - 1 < len(row) else None
        txt = ("" if cell is None else str(cell)).replace("\u00a0", " ").strip().lower()
        txt_norm = _re.sub(r"\s+", " ", txt)
        for canonical, patterns in compiled.items():
            if any(p.fullmatch(txt_norm) for p in patterns):
                return _METRIC_KEYS.get(canonical)
        r -= 1
    return None

def _num_at(values, row_1b: int, col_1b: int) -> float | None:
    row = values[row_1b - 1] if 0 < row_1b <= len(values) else []
    v = row[col_1b - 1] if 0 < col_1b <= len(row) else None
    f = _to_float(v)
    return None if math.isnan(f) else float(f)

@metrics_bp.post("/ingest-from-sheet")
def ingest_from_sheet():
    """
    Body: { "url": "<share link>", "sheet": "bCAS (Q4 Adj)", "conn_id": "<optional>" }
    Reads whole sheet, totals per month, and upserts into portfolio_period_metrics.
    """
    data = request.get_json(force=True) or {}
    share_url = (data.get("url") or data.get("share_url") or "").strip()
    sheet = data.get("sheet") or "bCAS (Q4 Adj)"
    conn_id = data.get("conn_id")

    bearer = _resolve_bearer()
    if not bearer:
        return jsonify(error="Unauthorized: sign in with Microsoft or enable app token"), 401

    # Resolve workbook: url -> conn_id -> latest/shared
    try:
        if share_url:
            drive_id, item_id = open_excel_by_share_url(share_url, bearer)
        elif conn_id:
            uid = _current_user_id()
            if not uid:
                return jsonify(error="Unauthorized"), 401
            conn = SharePointConnection.query.filter_by(id=conn_id, user_id=uid).first()
            if not conn:
                return jsonify(error="Connection not found"), 404
            drive_id, item_id = conn.drive_id, conn.item_id
        else:
            conn = _latest_sharepoint_connection_for_user_or_shared()
            if not conn:
                return jsonify(error="No SharePoint connection or URL provided."), 400
            drive_id, item_id = conn.drive_id, conn.item_id
    except Exception as e:
        return jsonify(error=f"Failed to open workbook: {e}"), 400

    sid = None
    try:
        sid = create_session(drive_id, item_id, bearer, persist=False)
        # Pull values
        tried = _sheet_candidates(sheet) or [sheet]
        values = None; last_err = None
        for sn in tried:
            try:
                values = used_range_values(drive_id, item_id, sn, bearer, sid)
                sheet = sn
                break
            except Exception as e:
                last_err = e
        if values is None:
            return jsonify(error=f"Worksheet not found (tried: {tried}). Last error: {last_err}"), 400

        # 1) Find the totals block anchor and header dates
        lbl_end = _find_label_row(values, "Ending Balance") or 60
        hdr_row, date_cols = _find_header_row_and_date_columns(values, max_scan_rows=120, anchor_row=lbl_end)
        if not hdr_row:
            return jsonify(error="No header row with dates found near totals block"), 400

        # 2) TOTAL-ROW path (preferred)
        total_row = _find_total_row(values)
        totals_by_date = {}

        from calendar import monthrange

        if total_row:
            for col_1b, dt in date_cols.items():
                metric = _metric_for_column(values, hdr_row, col_1b, lookback_rows=10)
                if not metric:
                    continue
                val = _num_at(values, total_row, col_1b)
                me = date(dt.year, dt.month, monthrange(dt.year, dt.month)[1])
                rec = totals_by_date.get(me) or dict(beginning=None, ending=None, unrealized=None, realized=None, fees=None)
                rec[metric] = val
                totals_by_date[me] = rec
        else:
            # 3) Fallback to label-row intersection if no "Total" row text exists
            lbl_rows = {
                "Beginning Balance": _find_label_row(values, "Beginning Balance"),
                "Ending Balance": _find_label_row(values, "Ending Balance"),
                "Unrealized Gain/Loss": _find_label_row(values, "Unrealized Gain/Loss"),
                "Realized Gain/Loss": _find_label_row(values, "Realized Gain/Loss"),
                "Management Fees": _find_label_row(values, "Management Fees"),
            }
            if not lbl_rows["Ending Balance"]:
                return jsonify(error="Could not locate 'Ending Balance' (or 'Current Value') label row"), 400

            for col_1b, dt in date_cols.items():
                rec = dict(beginning=None, ending=None, unrealized=None, realized=None, fees=None)
                if lbl_rows.get("Beginning Balance"):
                    rec["beginning"] = _num_at(values, lbl_rows["Beginning Balance"], col_1b)
                rec["ending"] = _num_at(values, lbl_rows["Ending Balance"], col_1b)
                if lbl_rows.get("Unrealized Gain/Loss"):
                    rec["unrealized"] = _num_at(values, lbl_rows["Unrealized Gain/Loss"], col_1b)
                if lbl_rows.get("Realized Gain/Loss"):
                    rec["realized"] = _num_at(values, lbl_rows["Realized Gain/Loss"], col_1b)
                if lbl_rows.get("Management Fees"):
                    rec["fees"] = _num_at(values, lbl_rows["Management Fees"], col_1b)
                me = date(dt.year, dt.month, monthrange(dt.year, dt.month)[1])
                totals_by_date[me] = rec

        # 4) Beginning(M) = Ending(M-1) when beginning is blank
        for dt_key in sorted(totals_by_date.keys()):
            if totals_by_date[dt_key]["beginning"] is None:
                prevs = [d for d in totals_by_date.keys() if d < dt_key]
                if prevs:
                    totals_by_date[dt_key]["beginning"] = totals_by_date[max(prevs)]["ending"]

        # 5) Upsert rows
        upserted = []
        for dt_key, rec in sorted(totals_by_date.items()):
            stmt = sqlite_insert(PortfolioPeriodMetric).values(
                sheet=sheet,
                as_of_date=dt_key,
                beginning_balance=rec["beginning"],
                ending_balance=rec["ending"],
                unrealized_gain_loss=rec["unrealized"],
                realized_gain_loss=rec["realized"],
                management_fees=rec["fees"],
                source="sharepoint-live",
            )
            stmt = stmt.on_conflict_do_update(
                index_elements=["sheet", "as_of_date"],
                set_={
                    "beginning_balance": stmt.excluded.beginning_balance,
                    "ending_balance": stmt.excluded.ending_balance,
                    "unrealized_gain_loss": stmt.excluded.unrealized_gain_loss,
                    "realized_gain_loss": stmt.excluded.realized_gain_loss,
                    "management_fees": stmt.excluded.management_fees,
                    "updated_at": datetime.utcnow(),
                },
            )
            db.session.execute(stmt)
            upserted.append(dt_key.isoformat())
        db.session.commit()

        return jsonify(ok=True, sheet=sheet, upserted=upserted)
    except Exception as e:
        logging.exception("ingest-from-sheet failed")
        return jsonify(error=str(e)), 500
    finally:
        if sid:
            try: close_session(drive_id, item_id, bearer, sid)
            except Exception: pass

@metrics_bp.get("/periods")
def list_periods():
    """
    Return normalized monthly totals for a sheet.
    """
    rows = (PortfolioPeriodMetric.query
            .filter_by(sheet=(request.args.get("sheet") or "bCAS (Q4 Adj)"))
            .order_by(PortfolioPeriodMetric.as_of_date.asc())
            .all())
    return jsonify([r.to_dict() for r in rows])

# =========================
#   MARKET endpoints
# =========================
@market_bp.post("/store_history")
def store_history():
    if upsert_history is None:
        return {"error": "market_store service not available"}, 500
    symbol = request.args.get("symbol", "^GSPC")
    start = request.args.get("start")
    end = request.args.get("end")
    interval = request.args.get("interval", "1d")
    if not start or not end:
        return {"error": "start and end are required (YYYY-MM-DD)"}, 400
    n = upsert_history(symbol, start, end, interval)
    return {"symbol": symbol, "inserted_or_updated": n, "status": "ok"}

@market_bp.get("/history_db")
def history_db():
    symbol = request.args.get("symbol", "^GSPC")
    start = request.args.get("start")
    end = request.args.get("end")
    q = MarketPrice.query.filter(MarketPrice.symbol == symbol)
    if start: q = q.filter(MarketPrice.date >= start)
    if end:   q = q.filter(MarketPrice.date <= end)
    rows = q.order_by(MarketPrice.date.asc()).all()
    return {
        "symbol": symbol,
        "rows": [
            {"date": r.date.isoformat(), "open": r.open, "high": r.high, "low": r.low,
             "close": r.close, "adj_close": r.adj_close, "volume": r.volume}
            for r in rows
        ]
    }


# === ALLOCATION (donut) ======================================================
@metrics_bp.get("/allocation")
def allocation():
    """
    Returns the portfolio allocation by investment for a given period.

    Query:
      ?period_end=YYYY-MM | YYYY-MM-DD   (optional; defaults to latest available)
      ?sheet=Master                      (optional; currently unused because
                                          investment values are sheet-agnostic)
      ?min_pct=0.0                       (optional; group very small slices into 'Other')
    """
    try:
        # normalize requested period (month-end)
        pe = (request.args.get("period_end") or "").strip()
        if pe:
            try:
                if len(pe) == 7 and pe[4] == "-":
                    y, m = int(pe[:4]), int(pe[5:7])
                    from calendar import monthrange
                    as_of = date(y, m, monthrange(y, m)[1])
                else:
                    dt = datetime.fromisoformat(pe).date()
                    from calendar import monthrange
                    as_of = date(dt.year, dt.month, monthrange(dt.year, dt.month)[1])
            except Exception:
                return jsonify(error="invalid period_end"), 400
        else:
            # find the latest date that has values
            as_of = db.session.query(func.max(PortfolioInvestmentValue.as_of_date)).scalar()
            if not as_of:
                return jsonify(error="no investment values found"), 404

        # sum by investment
        rows = (
            db.session.query(
                PortfolioInvestmentValue.investment_id.label("id"),
                func.sum(PortfolioInvestmentValue.value).label("value")
            )
            .filter(PortfolioInvestmentValue.as_of_date == as_of)
            .group_by(PortfolioInvestmentValue.investment_id)
            .all()
        )

        if not rows:
            return jsonify(error="no values for requested month"), 404

        # load investment metadata (name, color)
        inv_ids = [r.id for r in rows]
        inv_map = {
            i.id: i for i in Investment.query.filter(Investment.id.in_(inv_ids)).all()
        }

        total = float(sum(r.value or 0 for r in rows)) or 0.0
        if total <= 0:
            return jsonify(error="total is zero"), 404

        # a pleasant fallback palette (matches your UI vibe)
        palette = [
            "#6366F1", "#10B981", "#60A5FA", "#F59E0B", "#EF4444",
            "#8B5CF6", "#14B8A6", "#22C55E", "#3B82F6", "#EAB308",
            "#F97316", "#EC4899", "#06B6D4", "#84CC16"
        ]

        def pick_color(idx, default):
            return default or palette[idx % len(palette)]

        items = []
        for idx, r in enumerate(rows):
            inv = inv_map.get(r.id)
            name = inv.name if inv else f"Investment {r.id}"
            color = pick_color(idx, (inv.color_hex if inv and inv.color_hex else None))
            value = float(r.value or 0.0)
            pct = (value / total) * 100.0
            items.append({
                "id": int(r.id),
                "name": name,
                "value": round(value, 2),
                "percent": round(pct, 4),
                "color": color,
            })

        # optional “other” bucket
        try:
            min_pct = float(request.args.get("min_pct", "0") or 0)
        except Exception:
            min_pct = 0.0
        if min_pct > 0:
            major, minor = [], []
            for it in items:
                (major if it["percent"] >= min_pct else minor).append(it)
            if minor:
                other_value = sum(x["value"] for x in minor)
                other_pct = (other_value / total) * 100.0
                major.append({
                    "id": -1, "name": "Other",
                    "value": round(other_value, 2),
                    "percent": round(other_pct, 4),
                    "color": "#CBD5E1",   # slate-300
                })
                items = major

        # sort biggest → smallest for nicer labeling
        items.sort(key=lambda x: x["value"], reverse=True)

        return jsonify({
            "as_of": as_of.isoformat(),
            "total": round(total, 2),
            "items": items,
        })
    except Exception as e:
        current_app.logger.exception("allocation failed")
        return jsonify(error=str(e)), 500