# backend/routes/excel_routes.py
from __future__ import annotations

from flask import Blueprint, request, jsonify
from werkzeug.utils import secure_filename
import os, math, traceback
from datetime import datetime, timedelta, date
from calendar import monthrange
from typing import Dict, Any, List, Tuple, Optional

from openpyxl import load_workbook
from sqlalchemy.exc import SQLAlchemyError
from sqlalchemy.dialects.postgresql import insert as pg_insert
from sqlalchemy.dialects.sqlite import insert as sqlite_insert
import re  # <-- used for filename year detection

from backend.models import (
    db,
    ExcelUploadHistory,
    PortfolioPeriodMetric,
    Investment,
    PortfolioInvestmentValue,
)

# Reuse the same investor ingest logic used by SharePoint so behavior matches
from backend.routes.investor_sync_routes import _ingest_investor_values

# NEW: embed workbook into the vector KB so chat can retrieve from it

# Optional lineage record
try:
    from backend.models import DataSource
except Exception:
    DataSource = None  # pragma: no cover

# Optional current_user (if your app uses flask-login)
try:
    from flask_login import current_user  # type: ignore
except Exception:
    current_user = None  # pragma: no cover

excel_bp = Blueprint("excel_bp", __name__, url_prefix="/excel")

UPLOAD_FOLDER = os.path.join(os.getcwd(), "uploads")
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

DATE_COL = "as_of_date"


# ---------------------- generic helpers ----------------------
def _dialect_insert():
    name = (db.engine.name or "").lower()
    return pg_insert if "postgre" in name else sqlite_insert

def _to_float_cell(v) -> float:
    if v is None or v == "" or str(v).strip() in {"—", "-", "–"}:
        return math.nan
    s = str(v).strip().replace(",", "").replace("$", "")
    if s.startswith("(") and s.endswith(")"):
        s = "-" + s[1:-1]
    try:
        return float(s)
    except Exception:
        return math.nan

def _candidate_date_formats():
    return (
        "%m/%d/%Y", "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%y", "%d-%b-%Y",
        "%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S", "%d-%b-%y",
        "%b-%y", "%b %y", "%b-%Y", "%b %Y"
    )

def _maybe_excel_serial(v) -> Optional[date]:
    try:
        fv = float(v)
        if 20000 < fv < 90000:
            base = datetime(1899, 12, 30)
            return (base + timedelta(days=int(fv))).date()
    except Exception:
        pass
    return None

def _looks_like_date(v) -> bool:
    if isinstance(v, (datetime, date)):
        return True
    if _maybe_excel_serial(v):
        return True
    s = str(v).strip().rstrip("Z")
    for fmt in _candidate_date_formats():
        try:
            datetime.strptime(s, fmt)
            return True
        except Exception:
            continue
    return False

def _parse_date_any(v) -> Optional[date]:
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    ser = _maybe_excel_serial(v)
    if ser:
        return ser
    s = str(v).strip().rstrip("Z")
    for fmt in _candidate_date_formats():
        try:
            return datetime.strptime(s, fmt).date()
        except Exception:
            continue
    return None

import regex as _re
def _clean_txt(x: str) -> str:
    return _re.sub(r"\s+", " ", (x or "")).strip().lower()

def _values_from_openpyxl(ws) -> list[list]:
    vals = []
    for r in ws.iter_rows(values_only=True):
        vals.append(list(r))
    return vals


# ---------------------- ADMIN TOTALS INGEST (existing) ----------------------
_LABEL_ALIASES = {
    "Beginning Balance": [
        r"^begin(ning)? balance$", r"^opening (nav|balance)$",
        r"^current period begin(ning)? balance$", r"^total begin(ning)? balance$",
        r"^total beginning balance$",
    ],
    "Ending Balance": [
        r"^ending balance$", r"^closing balance$", r"^current value$",
        r"^total ending balance$", r"^total current value$",
    ],
    "Unrealized Gain/Loss": [
        r"^unrealis?ed gain/?loss$", r"^unrealized pnl$",
        r"^total unrealis?ed gain/?loss$",
    ],
    "Realized Gain/Loss": [
        r"^realis?ed gain/?loss$", r"^realized pnl$", r"^total realis?ed gain/?loss$",
    ],
    "Management Fees": [
        r"^management fees?$", r"^mgmt fees?$", r"^total management fees?$",
    ],
}
_METRIC_KEYS = {
    "Beginning Balance": "beginning",
    "Ending Balance": "ending",
    "Unrealized Gain/Loss": "unrealized",
    "Realized Gain/Loss": "realized",
    "Management Fees": "fees",
}

def _find_header_row_and_date_columns(values: List[List], max_scan_rows: int = 200, anchor_row: Optional[int] = None) -> Tuple[Optional[int], Dict[int, date]]:
    candidates: List[Tuple[int, Dict[int, date]]] = []
    rows = len(values)
    cols = max((len(r) for r in values), default=0)
    for r in range(1, min(rows, max_scan_rows) + 1):
        local: Dict[int, date] = {}
        row = values[r - 1]
        for c in range(1, cols + 1):
            v = row[c - 1] if c - 1 < len(row) else None
            if _looks_like_date(v):
                d = _parse_date_any(v)
                if d:
                    local[c] = d
        if len(local) >= 2:
            candidates.append((r, local))
    if not candidates:
        return None, {}
    if anchor_row:
        above = [(r, m) for (r, m) in candidates if r <= anchor_row]
        if above:
            above.sort(key=lambda x: (anchor_row - x[0], -len(x[1])))
            return above[0]
    candidates.sort(key=lambda x: (-len(x[1]), x[0]))
    return candidates[0]

def _metric_for_column(values: List[List], header_row_1b: int, col_1b: int) -> Optional[str]:
    compiled = {name: [_re.compile(p, _re.I) for p in pats] for name, pats in _LABEL_ALIASES.items()}
    for d in range(0, 13):
        for r in (header_row_1b - d, header_row_1b + d):
            if r < 1:
                continue
            row = values[r - 1] if r - 1 < len(values) else []
            cell = row[col_1b - 1] if col_1b - 1 < len(row) else None
            txt = _clean_txt("" if cell is None else str(cell))
            for canonical, patterns in compiled.items():
                if any(p.fullmatch(txt) for p in patterns):
                    return _METRIC_KEYS.get(canonical)
    return None

def _next_metric_label_below(values: List[List], start_row_1b: int, all_label_rows: List[Optional[int]]) -> Optional[int]:
    below = [r for r in all_label_rows if r and r > start_row_1b]
    return min(below) if below else None

def _sum_investor_rows_ignore_total(
    values: List[List], start_label_row_1b: int, date_col_1b: int,
    stop_row_1b: Optional[int] = None, name_col_1b: int = 2,
    id_col_1b: int = 1, max_blank_streak: int = 50
) -> Optional[float]:
    total = 0.0; have = False; blanks = 0
    FORBIDDEN = {"total", "entity level", "grand total"}
    r = start_label_row_1b + 1
    rows = len(values)
    limit = stop_row_1b if stop_row_1b else rows + 1
    while r < limit and r <= rows:
        row = values[r - 1] if r - 1 < len(values) else []
        name_val = row[name_col_1b - 1] if name_col_1b - 1 < len(row) else None
        id_val   = row[id_col_1b   - 1] if id_col_1b   - 1 < len(row) else None
        name_txt = _clean_txt("" if name_val is None else str(name_val))
        if name_txt in FORBIDDEN:
            r += 1; blanks = 0; continue
        is_partner_row = bool(name_txt) or (isinstance(id_val, (int, float)) or (isinstance(id_val, str) and id_val.strip()))
        if not is_partner_row:
            blanks += 1
            if blanks >= max_blank_streak: break
            r += 1; continue
        else:
            blanks = 0
        v = row[date_col_1b - 1] if date_col_1b - 1 < len(row) else None
        f = _to_float_cell(v)
        if not math.isnan(f):
            total += float(f); have = True
        r += 1
    return total if have else None

def _ingest_local_admin_totals(xlsx_path: str, sheet: str | None = None) -> dict:
    """
    Parses totals (Beginning/Ending/Unrealized/Realized/Fee) by month and upserts into PortfolioPeriodMetric.
    Returns: {"sheet": <resolved>, "upserted": [ISO dates], "values": <full used-range as lists>}
    """
    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    try:
        target = sheet or wb.sheetnames[0]
        if target not in wb.sheetnames:
            want = _clean_txt(target)
            for nm in wb.sheetnames:
                if _clean_txt(nm) == want or want in _clean_txt(nm):
                    target = nm; break

        ws = wb[target]
        values = _values_from_openpyxl(ws)

        # anchor near 'ending balance'
        lbl_end = None
        for r_i, row in enumerate(values[:200], start=1):
            if any("ending balance" in str(x).lower() for x in row if x is not None):
                lbl_end = r_i; break

        hdr_row, date_cols = _find_header_row_and_date_columns(values, max_scan_rows=200, anchor_row=(lbl_end or 200))
        if not hdr_row or not date_cols:
            raise RuntimeError("No header row with month/date columns found near totals block.")

        def _find_label(label):
            for r_i, row in enumerate(values[:250], start=1):
                if any(_clean_txt(str(x)) == _clean_txt(label) for x in row if x is not None):
                    return r_i
            return None

        lbl_begin = _find_label("Beginning Balance")
        lbl_end   = _find_label("Ending Balance")
        lbl_unrl  = _find_label("Unrealized Gain/Loss")
        lbl_rlzd  = _find_label("Realized Gain/Loss")
        lbl_fees  = _find_label("Management Fees")
        label_rows = {"beginning": lbl_begin, "ending": lbl_end, "unrealized": lbl_unrl, "realized": lbl_rlzd, "fees": lbl_fees}
        all_label_rows = [lbl_begin, lbl_end, lbl_unrl, lbl_rlzd, lbl_fees]

        totals_by_date: Dict[date, Dict[str, Optional[float]]] = {}
        for col1b, dt in sorted(date_cols.items()):
            as_of = date(dt.year, dt.month, monthrange(dt.year, dt.month)[1])
            rec = totals_by_date.get(as_of) or dict(beginning=None, ending=None, unrealized=None, realized=None, fees=None)

            metric = _metric_for_column(values, hdr_row, col1b)
            if not metric:
                totals_by_date[as_of] = rec; continue

            start_row = label_rows.get(metric)
            if not start_row:
                totals_by_date[as_of] = rec; continue

            stop_row = _next_metric_label_below(values, start_row, all_label_rows)
            val = _sum_investor_rows_ignore_total(values, start_label_row_1b=start_row, date_col_1b=col1b, stop_row_1b=stop_row)
            rec[metric] = val
            totals_by_date[as_of] = rec

        # carry forward missing beginnings
        for d in sorted(totals_by_date.keys()):
            if totals_by_date[d]["beginning"] is None:
                prevs = [p for p in totals_by_date if p < d]
                if prevs:
                    totals_by_date[d]["beginning"] = totals_by_date[max(prevs)]["ending"]

        insert_fn = _dialect_insert()
        upserted = []
        for dt_key, rec in sorted(totals_by_date.items()):
            values_row = dict(
                sheet=target,
                beginning_balance=rec["beginning"],
                ending_balance=rec["ending"],
                unrealized_gain_loss=rec["unrealized"],
                realized_gain_loss=rec["realized"],
                management_fees=rec["fees"],
                source="upload"
            )
            values_row[DATE_COL] = dt_key

            stmt = insert_fn(PortfolioPeriodMetric).values(**values_row)
            stmt = stmt.on_conflict_do_update(
                index_elements=["sheet", DATE_COL],
                set_={
                    "beginning_balance": stmt.excluded.beginning_balance,
                    "ending_balance": stmt.excluded.ending_balance,
                    "unrealized_gain_loss": stmt.excluded.unrealized_gain_loss,
                    "realized_gain_loss": stmt.excluded.realized_gain_loss,
                    "management_fees": stmt.excluded.management_fees,
                    "updated_at": datetime.utcnow(),
                    "source": stmt.excluded.source,
                },
            )
            db.session.execute(stmt)
            upserted.append(dt_key.isoformat())
        db.session.commit()

        if not upserted:
            raise RuntimeError("No month columns were detected in the uploaded sheet.")

        return {"sheet": target, "upserted": upserted, "values": values}
    finally:
        try: wb.close()
        except Exception: pass


# ---------------------- SHEET RESOLUTION (mirror SP) ----------------------
import re as _re2
def _normalize_sheet_name(s: str) -> str:
    return _re2.sub(r"[\s\-\_\.\(\)\[\]\{\}\+]+", "", (s or "").strip()).lower()

def _sheet_candidates(name: str) -> list[str]:
    base = (name or "").strip()
    if not base: return []
    variants = {
        base,
        base.replace("(", " (").replace("  ", " ").strip(),
        _re2.sub(r"\s*\(", " (", base).strip(),
        _re2.sub(r"\s+", "", base),
        base.replace("+", " "),
    }
    return [v for v in variants if v]


# ---------------------- FILE CLASSIFIER ----------------------
def _has_investments_table(values: List[List]) -> bool:
    limit = min(200, len(values))
    for r in range(limit):
        row = values[r] or []
        for cell in row:
            if str(cell or "").strip().lower() == "investments":
                return True
    return False

def _has_balance_labels(values: List[List]) -> bool:
    labels = {"beginning balance", "ending balance", "unrealized gain/loss", "realized gain/loss", "management fees"}
    limit = min(250, len(values))
    for r in range(limit):
        row = values[r] or []
        for cell in row:
            txt = str(cell or "").strip().lower()
            if txt in labels:
                return True
    return False

def _classify_workbook(values: List[List]) -> str:
    has_inv = _has_investments_table(values)
    has_bal = _has_balance_labels(values)
    if has_inv and has_bal: return "mixed"
    if has_inv: return "investment"
    if has_bal: return "balance"
    return "unknown"


# ---------------------- INVESTMENTS INGEST ----------------------
_MONTHS = {
    "jan": 1, "january": 1,
    "feb": 2, "february": 2,
    "mar": 3, "march": 3,
    "apr": 4, "april": 4,
    "may": 5,
    "jun": 6, "june": 6,
    "jul": 7, "july": 7,
    "aug": 8, "august": 8,
    "sep": 9, "sept": 9, "september": 9,
    "oct": 10, "october": 10,
    "nov": 11, "november": 11,
    "dec": 12, "december": 12,
}

def _find_header_row(values: List[List]) -> int:
    upto = min(80, len(values))
    for i in range(upto):
        row = [str(x or "").strip().lower() for x in values[i]]
        if any(_re.match(r"^invest(ment|ments)\b", c) for c in row):
            return i
    for i in range(upto):
        row = [str(x or "").strip().lower() for x in values[i]]
        if any(row):
            return i
    return 0

def _month_end(d: date) -> date:
    return date(d.year, d.month, monthrange(d.year, d.month)[1])

def _detect_year_banners(values: List[List], top_rows: int = 6) -> Dict[int, int]:
    year_by_col: Dict[int, int] = {}
    upto = min(top_rows, len(values))
    for r in range(upto):
        row = values[r] or []
        for j, cell in enumerate(row):
            if isinstance(cell, (int, float)):
                continue
            txt_raw = "" if cell is None else str(cell)
            txt = _clean_txt(txt_raw)
            if not txt:
                continue
            if _re.fullmatch(r"\s*(20\d{2})\s*", txt):
                year_by_col.setdefault(j, int(_re.fullmatch(r"\s*(20\d{2})\s*", txt).group(1)))
                continue
            if not _re.search(r"[a-z]", txt):
                continue
            m = _re.search(r"\b(20\d{2})\b", txt)
            if m:
                year_by_col.setdefault(j, int(m.group(1)))
    return year_by_col

def _detect_date_columns(
    values: List[List],
    header_row_idx: int,
    preferred_year: Optional[int] = None
) -> Dict[int, date]:
    from calendar import monthrange as _mr

    rows = len(values)
    cols = max((len(r) for r in values), default=0)

    def _row_txts(ridx: int) -> List[str]:
        row = values[ridx] if 0 <= ridx < rows else []
        return [_clean_txt(str(row[j])) if j < len(row) else "" for j in range(cols)]

    MONTHS = {
        "jan": 1, "january": 1, "feb": 2, "february": 2, "mar": 3, "march": 3,
        "apr": 4, "april": 4, "may": 5, "jun": 6, "june": 6, "jul": 7, "july": 7,
        "aug": 8, "august": 8, "sep": 9, "sept": 9, "september": 9,
        "oct": 10, "october": 10, "nov": 11, "november": 11,
        "dec": 12, "december": 12,
    }

    def _month_map_for_row(ridx: int) -> Dict[int, date]:
        out: Dict[int, date] = {}
        txts = _row_txts(ridx)
        for j, t in enumerate(txts):
            if t in MONTHS:
                m = MONTHS[t]
                y = preferred_year if preferred_year is not None else None
                if y:
                    out[j] = date(y, m, _mr(y, m)[1])
        return out

    if preferred_year is not None:
        out = _month_map_for_row(header_row_idx)
        if out:
            return out
        for off in [1, -1, 2, -2, 3, -3]:
            out = _month_map_for_row(header_row_idx + off)
            if out:
                return out
        for ridx in range(min(rows, 8)):
            out = _month_map_for_row(ridx)
            if out:
                return out

    for r in range(0, min(rows, 120)):
        row = values[r] if r < rows else []
        if any(_clean_txt(str(c)) == "ending date" for c in row):
            out: Dict[int, date] = {}
            for j, c in enumerate(row):
                dt = _parse_date_any(c)
                if dt:
                    out[j] = date(dt.year, dt.month, _mr(dt.year, dt.month)[1])
            if out:
                return out

    best_r = None
    best_count = 0
    for r in range(0, min(rows, 120)):
        row = values[r] if r < rows else []
        cnt = sum(1 for c in row if _parse_date_any(c))
        if cnt > best_count:
            best_count, best_r = cnt, r
    if best_r is not None and best_count >= 4:
        row = values[best_r]
        out: Dict[int, date] = {}
        for j, c in enumerate(row):
            dt = _parse_date_any(c)
            if dt:
                out[j] = date(dt.year, dt.month, _mr(dt.year, dt.month)[1])
        if out:
            return out

    header = values[header_row_idx] if header_row_idx < rows else []
    out: Dict[int, date] = {}
    for j in range(cols):
        dt = _parse_date_any(header[j] if j < len(header) else None)
        if dt:
            out[j] = date(dt.year, dt.month, _mr(dt.year, dt.month)[1])
    if out:
        return out

    year_banners: Dict[int, int] = _detect_year_banners(values, top_rows=6)

    def _nearest_banner_year_to_right(col: int) -> Optional[int]:
        rights = [c for c in sorted(year_banners) if c > col]
        return year_banners[rights[0]] if rights else None

    header_txts = _row_txts(header_row_idx)
    current_year: Optional[int] = None
    out = {}
    for j in range(cols):
        if j in year_banners:
            current_year = year_banners[j]
        hdr = header_txts[j]
        if hdr in MONTHS:
            y = current_year if current_year is not None else _nearest_banner_year_to_right(j)
            if y:
                m = MONTHS[hdr]
                out[j] = date(y, m, _mr(y, m)[1])
    if out:
        return out

    out = {}
    for j in range(cols):
        r = header_row_idx - 1
        while r >= 0:
            dt = _parse_date_any(values[r][j] if j < len(values[r]) else None)
            if dt:
                out[j] = date(dt.year, dt.month, _mr(dt.year, dt.month)[1])
                break
            r -= 1
    return out


def _ensure_color(i: int) -> str:
    palette = [
        "#6366F1", "#10B981", "#F59E0B", "#EC4899", "#14B8A6",
        "#F43F5E", "#22C55E", "#8B5CF6", "#06B6D4", "#F97316",
        "#84CC16", "#3B82F6", "#A855F7", "#0EA5E9", "#D946EF",
    ]
    return palette[i % len(palette)]

def _ingest_investments_table(
    values: List[List],
    resolved_sheet: str,
    source_id: Optional[int],
    preferred_year: Optional[int] = None,
) -> Dict[str, Any]:
    if not values or len(values) < 2:
        return {"ok": False, "error": "No cells to parse", "sheet": resolved_sheet}

    header_row_idx = _find_header_row(values)
    header = [str(c or "").strip() for c in values[header_row_idx]]
    header_lc = [h.lower() for h in header]

    name_idx = None
    for j, h in enumerate(header_lc):
        if _re.match(r"^invest(ment|ments)\b", h or ""):
            name_idx = j
            break
    if name_idx is None:
        return {
            "ok": False,
            "investments": 0,
            "values": 0,
            "sheet": resolved_sheet,
            "note": "No Investments column found in header.",
        }

    date_cols = _detect_date_columns(values, header_row_idx, preferred_year=preferred_year)
    if not date_cols:
        return {"ok": False, "error": "Could not locate monthly date columns for investments grid.", "sheet": resolved_sheet}

    if preferred_year is not None:
        date_cols = {j: d for j, d in date_cols.items() if d.year == preferred_year}
        if not date_cols:
            return {"ok": False, "error": f"No monthly columns resolved for year {preferred_year}.", "sheet": resolved_sheet}

    body = values[header_row_idx + 1 :]

    inserted_vals = 0
    ensured: Dict[str, int] = {}
    ensured_order: List[str] = []

    STOP_AT_NAME = "portfolio total"
    blank_streak = 0

    for r in body:
        name = str(r[name_idx] if name_idx < len(r) else "").strip()
        clean_name = _clean_txt(name)

        if clean_name == STOP_AT_NAME:
            break

        if not name:
            blank_streak += 1
            if blank_streak > 20:
                break
            continue
        blank_streak = 0
        if clean_name in {"total", "grand total"}:
            continue

        if name not in ensured:
            inv = Investment.query.filter_by(name=name).first()
            if not inv:
                inv = Investment(name=name, color_hex=_ensure_color(len(ensured_order)))
                db.session.add(inv)
                db.session.flush()
            ensured[name] = int(inv.id)
            ensured_order.append(name)
        inv_id = ensured[name]

        for j0b, mdt in sorted(date_cols.items(), key=lambda x: (x[1].year, x[1].month)):
            v = r[j0b] if j0b < len(r) else None
            f = _to_float_cell(v)
            if math.isnan(f):
                continue

            row = PortfolioInvestmentValue.query.filter_by(investment_id=inv_id, as_of_date=mdt).first()
            if row is None:
                row = PortfolioInvestmentValue(investment_id=inv_id, as_of_date=mdt)
            row.value = float(f)
            row.source = "valuation_sheet"
            row.source_id = source_id
            db.session.add(row)
            inserted_vals += 1

    db.session.commit()
    return {
        "ok": True,
        "investments": len(ensured),
        "values": inserted_vals,
        "sheet": resolved_sheet,
        "year": preferred_year,
    }


# ---------------------- ROUTE: upload + (admin + investor + investments + KB embed) ----------------------
@excel_bp.post("/upload_and_ingest")
def upload_and_ingest():
    """
    Form-data:
      - file: Excel (.xlsx/.xlsm/.xls)
      - sheet: optional worksheet name
    """
    f = request.files.get("file")
    if not f:
        return jsonify(error="No file uploaded"), 400

    sheet = (request.form.get("sheet") or request.args.get("sheet") or "").strip()
    filename = secure_filename(f.filename or "")
    if not filename:
        return jsonify(error="Invalid filename"), 400

    # robust year parse (works with underscores/dashes/spaces)
    m = re.search(r"(?<!\d)(20\d{2})(?!\d)", filename)
    preferred_year = int(m.group(1)) if m else None

    ext = os.path.splitext(filename)[1].lower()
    if ext not in {".xlsx", ".xlsm", ".xls"}:
        return jsonify(error=f"Unsupported type: {ext}"), 400

    path = os.path.join(UPLOAD_FOLDER, filename)
    f.save(path)

    try:
        # --- Resolve sheet + read values FIRST (so we can classify before ingesting) ---
        wb = load_workbook(path, data_only=True, read_only=True)
        try:
            if sheet and sheet in wb.sheetnames:
                resolved_sheet = sheet
            else:
                resolved_sheet = None
                if sheet:
                    for sn in (_sheet_candidates(sheet) or [sheet]):
                        if sn in wb.sheetnames:
                            resolved_sheet = sn; break
                    if not resolved_sheet:
                        want = _normalize_sheet_name(sheet)
                        for n in wb.sheetnames:
                            if _normalize_sheet_name(n) == want:
                                resolved_sheet = n; break
                        if not resolved_sheet:
                            for n in wb.sheetnames:
                                if want in _normalize_sheet_name(n):
                                    resolved_sheet = n; break
                if not resolved_sheet:
                    resolved_sheet = "Master" if "Master" in wb.sheetnames else wb.sheetnames[0]

            ws = wb[resolved_sheet]
            values = [[c.value for c in row]
                      for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column)]
        finally:
            wb.close()

        # --- Classify THIS sheet and decide which pipelines to run ---
        normalized = _normalize_sheet_name(resolved_sheet)
        file_type = _classify_workbook(values or [])  # investment / balance / mixed / unknown

        is_master = normalized == "master"
        looks_balance = ("bcas" in normalized) or ("q4adj" in normalized) or (file_type == "balance")
        looks_invest  = is_master or (file_type == "investment")

        if file_type == "mixed":
            looks_invest = is_master
            looks_balance = not is_master

        # --- Optional lineage record (before any writes that need ds_id) ---
        ds_id = None
        if DataSource is not None:
            try:
                added_by = None
                if current_user and getattr(current_user, "is_authenticated", False):
                    added_by = getattr(current_user, "email", None) or getattr(current_user, "username", None)
                ds = DataSource(
                    kind="upload",
                    file_name=filename,
                    sheet_name=resolved_sheet,
                    added_by=added_by
                )
                db.session.add(ds)
                db.session.flush()
                ds_id = ds.id
                db.session.commit()
            except Exception:
                db.session.rollback()
                ds_id = None  # non-fatal

        # --- 1) Admin totals ingest (ONLY for balance-type sheets) ---
        admin_res = {"upserted": []}
        if looks_balance:
            admin_res = _ingest_local_admin_totals(path, resolved_sheet)

        # --- 2) Investor ingest (ONLY for balance-type sheets) ---
        payload = {}
        if looks_balance:
            sp_resp, sp_status = _ingest_investor_values(
                values,
                resolved_sheet,
                drive_id=None,
                item_id=None,
                source="upload",
            )
            payload = sp_resp.get_json() if hasattr(sp_resp, "get_json") else {}

        # --- 3) Investments ingest (ONLY for investment-type sheets, e.g., 'Master') ---
        inv_payload = {}
        if looks_invest:
            try:
                inv_payload = _ingest_investments_table(
                    values or [], resolved_sheet, ds_id, preferred_year=preferred_year
                )
            except Exception as e:
                inv_payload = {"ok": False, "error": f"Investments ingest failed: {e}", "sheet": resolved_sheet}


        # --- Record upload history & cleanup ---
        db.session.add(ExcelUploadHistory(filename=filename, uploaded_at=datetime.utcnow()))
        db.session.commit()
        try: os.remove(path)
        except Exception: pass

        return jsonify({
            "ok": True,
            "sheet": resolved_sheet,
            "file_type": file_type,
            "data_source_id": int(ds_id) if ds_id is not None else None,
            "admin_periods_upserted": admin_res["upserted"],
            "investments_result": inv_payload,
            **payload,
        }), 200

    except (RuntimeError, SQLAlchemyError) as e:
        db.session.rollback()
        traceback.print_exc()
        return jsonify(error=f"Upload/ingest failed: {e}"), 500
    except Exception as e:
        db.session.rollback()
        traceback.print_exc()
        return jsonify(error="Upload/ingest failed. See server logs."), 500
