from __future__ import annotations
import io, os, json, time, hashlib
from dataclasses import dataclass
from typing import Any, Dict, List, Optional, Tuple
import numpy as np
from openpyxl import load_workbook

from backend.services.openai_client import LLMClient

# ---------------- Config (overridable via env) ----------------
EMBED_CACHE_DIR = os.environ.get("EMBED_CACHE_DIR", "backend/.embed_cache")
os.makedirs(EMBED_CACHE_DIR, exist_ok=True)

# How many records to embed per API call (keeps token count below request caps)
EMBED_BATCH_SIZE = int(os.environ.get("EMBED_BATCH_SIZE", "128"))

# Sheets/rows/columns limits (set to 0 to mean "no limit")
EMBED_SHEETS_LIMIT = int(os.environ.get("EMBED_SHEETS_LIMIT", "0"))          # 0 = all sheets
EMBED_MAX_ROWS_PER_SHEET = int(os.environ.get("EMBED_MAX_ROWS_PER_SHEET", "0"))  # 0 = all rows

# Per-record compaction to keep each text compact and cheap
EMBED_MAX_COLS = int(os.environ.get("EMBED_MAX_COLS", "80"))                # cap columns per record
EMBED_MAX_NUMERIC_COLS = int(os.environ.get("EMBED_MAX_NUMERIC_COLS", "32"))
EMBED_MAX_NONNUM_COLS = int(os.environ.get("EMBED_MAX_NONNUM_COLS", "16"))
EMBED_MAX_CELL_CHARS = int(os.environ.get("EMBED_MAX_CELL_CHARS", "160"))
EMBED_MAX_RECORD_CHARS = int(os.environ.get("EMBED_MAX_RECORD_CHARS", "2400"))

# Force rebuild (bypass cache) on every request (useful while tuning)
EMBED_REBUILD_ALWAYS = os.environ.get("EMBED_REBUILD_ALWAYS", "false").lower() == "true"

@dataclass
class RowRecord:
    sheet: str
    row: int
    label: str  # first column or synthetic label
    columns: Dict[str, Any]  # header -> value

def _cosine_sim(a: np.ndarray, b: np.ndarray) -> float:
    denom = (np.linalg.norm(a) * np.linalg.norm(b)) or 1e-8
    return float(np.dot(a, b) / denom)

def _hash_key(meta: Dict[str, Any]) -> str:
    j = json.dumps(meta, sort_keys=True, default=str)
    return hashlib.sha1(j.encode("utf-8")).hexdigest()

def _workbook_meta(local_path: Optional[str], provider_meta: Dict[str, Any]) -> Dict[str, Any]:
    meta = {}
    if local_path:
        ap = os.path.abspath(local_path)
        mtime = os.path.getmtime(ap) if os.path.exists(ap) else time.time()
        meta.update({"source": "upload", "path": ap, "mtime": mtime})
    else:
        meta.update(provider_meta or {})
    return meta

def _cache_paths(key: str) -> Tuple[str, str]:
    base = os.path.join(EMBED_CACHE_DIR, key)
    return base + ".json", base + ".npy"

def _load_cache(key: str) -> Optional[Tuple[Dict[str, Any], np.ndarray]]:
    jpath, npypath = _cache_paths(key)
    if not (os.path.exists(jpath) and os.path.exists(npypath)):
        return None
    with open(jpath, "r", encoding="utf-8") as f:
        meta = json.load(f)
    vecs = np.load(npypath)
    return meta, vecs

def _save_cache(key: str, meta: Dict[str, Any], vecs: np.ndarray) -> None:
    jpath, npypath = _cache_paths(key)
    with open(jpath, "w", encoding="utf-8") as f:
        json.dump(meta, f, ensure_ascii=False)
    np.save(npypath, vecs)

# ---------------- Workbook → Records ----------------

def _truncate(s: str, n: int) -> str:
    if len(s) <= n:
        return s
    return s[: max(0, n - 1)] + "…"

def _cell_str(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, float):
        return f"{v:.8g}"
    return str(v)

def _records_from_workbook(book_bytes: bytes) -> List[RowRecord]:
    """
    Index ALL sheets (unless EMBED_SHEETS_LIMIT > 0).
    Supports multi-row headers (first 2–3 rows) merged into a single header string.
    """
    wb = load_workbook(io.BytesIO(book_bytes), data_only=True, read_only=True)
    sheets = wb.sheetnames

    if EMBED_SHEETS_LIMIT and EMBED_SHEETS_LIMIT > 0 and len(sheets) > EMBED_SHEETS_LIMIT:
        sheets = sheets[:EMBED_SHEETS_LIMIT]

    records: List[RowRecord] = []

    def non_empty_count(row_cells):
        return sum(1 for c in row_cells if (c.value not in (None, "")))

    for sname in sheets:
        ws = wb[sname]
        if ws.max_row < 2 or ws.max_column < 1:
            continue

        # ---- multi-row header detection/merge (up to 3 rows) ----
        hdr_rows = [ws[i] for i in range(1, min(3, ws.max_row) + 1)]
        scores = [non_empty_count(r) for r in hdr_rows]
        base_idx = int(max(range(len(scores)), key=lambda i: scores[i]))
        base = hdr_rows[base_idx]

        merged_headers: List[str] = []
        for col_idx in range(1, ws.max_column + 1):
            parts = []
            if base_idx - 1 >= 0:
                v = hdr_rows[base_idx - 1][col_idx - 1].value
                if v not in (None, ""): parts.append(_truncate(str(v).strip(), EMBED_MAX_CELL_CHARS))
            bv = base[col_idx - 1].value
            if bv not in (None, ""): parts.append(_truncate(str(bv).strip(), EMBED_MAX_CELL_CHARS))
            if base_idx + 1 < len(hdr_rows):
                v = hdr_rows[base_idx + 1][col_idx - 1].value
                if v not in (None, ""): parts.append(_truncate(str(v).strip(), EMBED_MAX_CELL_CHARS))
            header = " | ".join([p for p in parts if p]).strip()
            if not header:
                header = f"Col{col_idx}"
            merged_headers.append(header)

        data_start_row = base_idx + 2
        if data_start_row > ws.max_row:
            continue

        # iterate rows (optionally capped)
        max_row = ws.max_row
        if EMBED_MAX_ROWS_PER_SHEET and EMBED_MAX_ROWS_PER_SHEET > 0:
            max_row = min(ws.max_row, data_start_row - 1 + EMBED_MAX_ROWS_PER_SHEET)

        row_num = data_start_row
        for r in ws.iter_rows(min_row=data_start_row, max_row=max_row, values_only=True):
            row_map: Dict[str, Any] = {}
            for idx, val in enumerate(r):
                if idx >= len(merged_headers):
                    break
                h = merged_headers[idx]
                if isinstance(val, (int, float)):
                    row_map[h] = val
                else:
                    row_map[h] = _truncate(_cell_str(val), EMBED_MAX_CELL_CHARS)
            if not any(v not in (None, "", 0) for v in row_map.values()):
                row_num += 1
                continue
            label_val = r[0] if len(r) > 0 else ""
            label = _truncate(_cell_str(label_val).strip(), 100) if label_val not in (None, "") else ""
            records.append(RowRecord(sheet=sname, row=row_num, label=label, columns=row_map))
            row_num += 1

    return records

def _build_texts(records: List[RowRecord]) -> List[str]:
    texts = []
    for rec in records:
        items = list(rec.columns.items())
        numeric = [(k, v) for k, v in items if isinstance(v, (int, float))][:EMBED_MAX_NUMERIC_COLS]
        nonnum = [(k, v) for k, v in items if not isinstance(v, (int, float))][:EMBED_MAX_NONNUM_COLS]
        pairs = [f"{k}: {v}" for k, v in (numeric + nonnum)]
        if EMBED_MAX_COLS and len(pairs) > EMBED_MAX_COLS:
            pairs = pairs[:EMBED_MAX_COLS]
        base = f"Sheet: {rec.sheet} | Row: {rec.row} | Label: {rec.label} | " + " | ".join(pairs)
        texts.append(_truncate(base, EMBED_MAX_RECORD_CHARS))
    return texts

# ---------------- Embedding index ----------------

class WorkbookIndex:
    def __init__(self, meta_key: str, records: List[RowRecord], embeddings: np.ndarray, dim: int):
        self.key = meta_key
        self.records = records
        self.embeddings = embeddings  # shape (N, D)
        self.dim = dim

    def topk(self, query_vec: np.ndarray, k: int = 5) -> List[Tuple[RowRecord, float]]:
        sims = [ _cosine_sim(e, query_vec) for e in self.embeddings ]
        idxs = np.argsort(sims)[::-1][:k]
        return [(self.records[i], sims[i]) for i in idxs]

def _embed_in_batches(llm: LLMClient, texts: List[str], batch_size: int) -> Tuple[np.ndarray, int]:
    if not texts:
        return np.zeros((0, 1), dtype=np.float32), 1
    vecs: List[List[float]] = []
    dim: Optional[int] = None
    n = len(texts)
    for i in range(0, n, batch_size):
        chunk = texts[i:i+batch_size]
        chunk_vecs = llm.embed(chunk)  # List[List[float]]
        if chunk_vecs and dim is None:
            dim = len(chunk_vecs[0])
        vecs.extend(chunk_vecs)
    return np.array(vecs, dtype=np.float32), (dim or 1536)

def build_or_load_index(
    llm: LLMClient,
    book_bytes: bytes,
    local_path: Optional[str] = None,
    provider_meta: Optional[Dict[str, Any]] = None,
) -> WorkbookIndex:
    records = _records_from_workbook(book_bytes)
    texts = _build_texts(records)

    meta = _workbook_meta(local_path, provider_meta or {})
    meta.update({
        "count": len(texts),
        "batch": EMBED_BATCH_SIZE,
        "cols": EMBED_MAX_COLS,
        "nnum": EMBED_MAX_NONNUM_COLS,
        "nnumc": EMBED_MAX_NUMERIC_COLS,
        "cell": EMBED_MAX_CELL_CHARS,
        "rec": EMBED_MAX_RECORD_CHARS,
        "rows_per_sheet": EMBED_MAX_ROWS_PER_SHEET,
        "sheets_limit": EMBED_SHEETS_LIMIT,
        "rebuild": EMBED_REBUILD_ALWAYS,
    })

    key = _hash_key(meta)
    if not EMBED_REBUILD_ALWAYS:
        cache = _load_cache(key)
        if cache:
            cached_meta, vecs = cache
            if cached_meta.get("count") == meta["count"]:
                return WorkbookIndex(key, records, vecs, vecs.shape[1] if vecs.ndim == 2 else 1536)

    vecs, dim = _embed_in_batches(llm, texts, EMBED_BATCH_SIZE)
    _save_cache(key, meta, vecs)
    return WorkbookIndex(key, records, vecs, dim)

def answer_from_topk(
    llm: LLMClient,
    query: str,
    index: WorkbookIndex,
    k: int = 5,
) -> Dict[str, Any]:
    qv = np.array(llm.embed([query])[0], dtype=np.float32)
    hits = index.topk(qv, k=k)

    context_lines = []
    citations = []
    for rec, score in hits:
        parts = [f"[{rec.sheet}] Row={rec.row} Label={rec.label} (sim={score:.2f})"]
        items = list(rec.columns.items())
        numeric = [(k, v) for k, v in items if isinstance(v, (int, float))][:6]
        nonnum = [(k, v) for k, v in items if not isinstance(v, (int, float))][:3]
        parts += [f"{k}={v}" for k, v in (numeric + nonnum) if v not in (None, "")]
        context_lines.append(" | ".join(parts))
        citations.append({"sheet": rec.sheet, "row": rec.row})

    prompt = (
        "You are given the user's question and the top-matching rows from an Excel workbook.\n"
        "Use ONLY the information in the context to answer. If the exact value is unclear, state the best match and explain briefly.\n"
        f"Question: {query}\n\nContext:\n- " + "\n- ".join(context_lines)
    )
    content = llm.finance_answer(prompt)
    return {"type": "nlp", "answer": content, "citations": citations, "source_rows": len(hits)}

# --------- Cache maintenance helpers (for endpoints) ---------

def clear_cache_for_path(local_path: Optional[str], provider_meta: Optional[Dict[str, Any]]) -> int:
    """
    Deletes cache files whose prefix matches the current meta (loose match).
    Returns number of files removed.
    """
    meta = _workbook_meta(local_path, provider_meta or {})
    prefix = _hash_key({k: meta[k] for k in sorted(meta.keys()) if k in ("path","source","drive_id","item_id")})
    removed = 0
    for f in os.listdir(EMBED_CACHE_DIR):
        if f.startswith(prefix):
            try:
                os.remove(os.path.join(EMBED_CACHE_DIR, f))
                removed += 1
            except Exception:
                pass
    return removed

def index_stats(llm: LLMClient, book_bytes: bytes, local_path: Optional[str], provider_meta: Optional[Dict[str, Any]]) -> Dict[str, Any]:
    recs = _records_from_workbook(book_bytes)
    return {
        "sheets": len(set(r.sheet for r in recs)),
        "records": len(recs),
        "example": _build_texts(recs[:2]),
        "path": local_path or provider_meta,
    }
